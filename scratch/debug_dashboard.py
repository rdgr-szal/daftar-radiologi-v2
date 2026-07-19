import openpyxl
import os
import datetime

filepath = "/Users/SZAL/Documents/Business/Web Dev/RADIOLOGY/Kesihatan Awam/Buku Daftar Radiologi (PER.SS-RA 101 Compliance)/Daftar_Radiologi/Pendaftaran/2026/07_JUL/2026 7JUL PER.SS-RA 101.xlsx"

def parse_mykad_py(ic):
    cleaned = "".join(filter(str.isdigit, ic))
    if len(cleaned) != 12:
        return None
    dob_part = cleaned[0:6]
    gender_digit = int(cleaned[11])
    return {
        "is_male": (gender_digit % 2 != 0)
    }

if os.path.exists(filepath):
    print("File exists!")
    wb = openpyxl.load_workbook(filepath, data_only=True)
    for sheetname in wb.sheetnames:
        if sheetname.isdigit():
            sheet = wb[sheetname]
            for r in range(9, 74):
                xray_no = sheet.cell(row=r, column=3).value
                if xray_no:
                    gender_cell = sheet.cell(row=r, column=7).value
                    ic = sheet.cell(row=r, column=4).value
                    name = sheet.cell(row=r, column=5).value
                    
                    g_str = ""
                    if gender_cell:
                        g_str = str(gender_cell).strip().upper()
                    
                    fallback = ""
                    if not g_str or g_str.startswith("="):
                        parsed = parse_mykad_py(str(ic))
                        if parsed:
                            g_str = "L" if parsed["is_male"] else "P"
                            fallback = " (Fallback from MyKad)"
                    
                    print(f"Row {r}: X-ray={xray_no}, IC={ic}, Name={name}, RawGender={gender_cell}, FinalGender={g_str}{fallback}")
else:
    print("File does not exist at:", filepath)
