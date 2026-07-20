import os
import sys
import json
import datetime
import shutil
from flask import Flask, render_template, request, jsonify, redirect, url_for
import openpyxl
import webbrowser
import threading

if getattr(sys, 'frozen', False):
    # Dijalankan sebagai executable PyInstaller
    template_folder = os.path.join(sys._MEIPASS, 'templates')
    static_folder = os.path.join(sys._MEIPASS, 'static')
    TEMPLATE_FILE_PATH = os.path.join(sys._MEIPASS, 'template.xlsx')
    app = Flask(__name__, template_folder=template_folder, static_folder=static_folder)
    # Folder data Excel berada bersebelahan dengan fail executable (.exe / App)
    BASE_DIR = os.path.dirname(sys.executable)
else:
    # Dijalankan sebagai kod sumber biasa
    app = Flask(__name__)
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    TEMPLATE_FILE_PATH = os.path.join(BASE_DIR, 'template.xlsx')

PENDAFTARAN_DIR = os.path.join(BASE_DIR, "Pendaftaran")
EXTRA_DATA_PATH = os.path.join(PENDAFTARAN_DIR, "extra_phris_data.json")
CONFIG_PATH = os.path.join(PENDAFTARAN_DIR, "config.json")

# Pemetaan Bulan dalam Bahasa Melayu
MONTH_MAP = {
    1: ("01_JAN", "1JAN"),
    2: ("02_FEB", "2FEB"),
    3: ("03_MAC", "3MAC"),
    4: ("04_APR", "4APR"),
    5: ("05_MEI", "5MEI"),
    6: ("06_JUN", "6JUN"),
    7: ("07_JUL", "7JUL"),
    8: ("08_OGOS", "8OGOS"),
    9: ("09_SEPT", "9SEPT"),
    10: ("10_OKT", "10OKT"),
    11: ("11_NOV", "11NOV"),
    12: ("12_DIS", "12DIS"),
}

# Utiliti: Mengambil jalan fail Excel berdasarkan tarikh
def get_excel_path(date_obj):
    year = date_obj.year
    month = date_obj.month
    if month not in MONTH_MAP:
        return None
    folder_name, file_month = MONTH_MAP[month]
    
    # Memuatkan singkatan daripada konfigurasi JSON
    config = load_config()
    singkatan = config.get("singkatan_klinik", "KKRP").upper()
    
    # 1. Cuba nama fail yang mengandungi singkatan klinik
    filename = f"{year} {file_month} PER.SS-RA 101 {singkatan}.xlsx"
    path = os.path.join(PENDAFTARAN_DIR, str(year), folder_name, filename)
    if os.path.exists(path):
        return path
        
    # 2. Cuba nama fail tanpa singkatan klinik (lalai)
    filename_no_suffix = f"{year} {file_month} PER.SS-RA 101.xlsx"
    path_no_suffix = os.path.join(PENDAFTARAN_DIR, str(year), folder_name, filename_no_suffix)
    if os.path.exists(path_no_suffix):
        return path_no_suffix
        
    # 3. Cuba cari mana-mana fail Excel pendaftaran di dalam direktori berkenaan
    dir_path = os.path.join(PENDAFTARAN_DIR, str(year), folder_name)
    if os.path.exists(dir_path):
        try:
            for f in os.listdir(dir_path):
                if "PER.SS-RA 101" in f and f.endswith(".xlsx") and not f.startswith("~$"):
                    return os.path.join(dir_path, f)
        except Exception:
            pass
            
    return path  # Fallback ke path asal jika semua di atas tidak wujud

# Utiliti: Load / Save Extra PHRIS Data (JSON Sidecar)
def load_extra_phris_data():
    if not os.path.exists(EXTRA_DATA_PATH):
        return {}
    try:
        with open(EXTRA_DATA_PATH, "r") as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading extra PHRIS data: {e}")
        return {}

def save_extra_phris_data(data):
    try:
        with open(EXTRA_DATA_PATH, "w") as f:
            json.dump(data, f, indent=4)
        return True
    except Exception as e:
        print(f"Error saving extra PHRIS data: {e}")
        return False

# Utiliti: Load / Save Config Data (JSON)
def load_config():
    if not os.path.exists(CONFIG_PATH):
        return {
            "is_configured": False,
            "klinik_asal": "KK Bandar Baru Selayang",
            "klinik_rujukan": [
                "KK Selayang Baru",
                "KK Gombak Setia",
                "KK Taman Ehsan"
            ]
        }
    try:
        with open(CONFIG_PATH, "r") as f:
            return json.load(f)
    except Exception as e:
        print(f"Error loading config: {e}")
        return {
            "is_configured": False,
            "klinik_asal": "KK Bandar Baru Selayang",
            "klinik_rujukan": ["KK Selayang Baru", "KK Gombak Setia", "KK Taman Ehsan"]
        }

def save_config(config_data):
    try:
        with open(CONFIG_PATH, "w") as f:
            json.dump(config_data, f, indent=4)
        return True
    except Exception as e:
        print(f"Error saving config: {e}")
        return False

# Utiliti: Sinkronisasi senarai klinik rujukan ke helaian Masterlist semua fail Excel
def sync_excel_masterlists(referral_clinics, klinik_asal, singkatan_klinik):
    import gc
    if not os.path.exists(PENDAFTARAN_DIR):
        return False
    
    # 1. Nama semula semua fail Excel dahulu mengikut singkatan fasiliti baru
    for root, dirs, files in os.walk(PENDAFTARAN_DIR):
        for file in files:
            if "PER.SS-RA 101" in file and file.endswith(".xlsx") and not file.startswith("~$"):
                # Sokong DUA format nama fail:
                # Lama (tanpa singkatan): "2026 1JAN PER.SS-RA 101.xlsx"
                # Ada singkatan lama:     "2026 1JAN PER.SS-RA 101 KKRP.xlsx"
                prefix = file.split(" PER.SS-RA 101")[0]
                new_filename = f"{prefix} PER.SS-RA 101 {singkatan_klinik}.xlsx"
                if file != new_filename:
                    old_filepath = os.path.join(root, file)
                    new_filepath = os.path.join(root, new_filename)
                    try:
                        os.rename(old_filepath, new_filepath)
                        print(f"Menamakan semula: {file} -> {new_filename}")
                    except Exception as e:
                        print(f"Ralat menamakan semula {file}: {e}")
                            
    # 2. Kemas kini sel-sel di dalam semua fail Excel yang telah dinamakan semula
    updated_files = 0
    referrals_only = [c for c in referral_clinics if c != klinik_asal]
    num_ref = len(referrals_only)
    ref_rows_count = max(8, num_ref)
    
    for root, dirs, files in os.walk(PENDAFTARAN_DIR):
        for file in files:
            if file.endswith(".xlsx") and not file.startswith("~$"):
                filepath = os.path.join(root, file)
                try:
                    wb = openpyxl.load_workbook(filepath)
                    
                    # A. Kemas kini Masterlist (C4 dan ke bawah)
                    if "Masterlist" in wb.sheetnames:
                        sheet = wb["Masterlist"]
                        sheet.cell(row=3, column=3).value = "KLINIK"
                        for r_idx in range(4, 100):
                            sheet.cell(row=r_idx, column=3).value = None
                        for idx, clinic_name in enumerate(referral_clinics):
                            sheet.cell(row=4 + idx, column=3).value = clinic_name
                                
                    # B. Kemas kini Data Penuh
                    if "Data Penuh" in wb.sheetnames:
                        sheet = wb["Data Penuh"]
                        sheet.cell(row=13, column=1).value = "KLINIK KESIHATAN"
                        
                        # Kosongkan baris 14 hingga 60 untuk kolum A dahulu
                        for r in range(14, 60):
                            sheet.cell(row=r, column=1).value = None
                            
                        # Tulis senarai klinik rujukan luar
                        for idx in range(ref_rows_count):
                            r_row = 14 + idx
                            if idx < num_ref:
                                sheet.cell(row=r_row, column=1).value = referrals_only[idx]
                            else:
                                sheet.cell(row=r_row, column=1).value = ""
                                
                        row_klinik_asal = 14 + ref_rows_count
                        sheet.cell(row=row_klinik_asal, column=1).value = klinik_asal
                        sheet.cell(row=row_klinik_asal + 1, column=1).value = "JUMLAH PESAKIT"
                        
                        # Tulis semua label tetap di bawah senarai klinik rujukan
                        sheet.cell(row=row_klinik_asal + 3, column=1).value = "TB"
                        sheet.cell(row=row_klinik_asal + 4, column=1).value = singkatan_klinik
                        sheet.cell(row=row_klinik_asal + 5, column=1).value = "KLINIK LUAR"
                        
                        sheet.cell(row=row_klinik_asal + 9, column=1).value = "WHEELCHAIR"
                        sheet.cell(row=row_klinik_asal + 10, column=1).value = "TROLLEY"
                        
                        sheet.cell(row=row_klinik_asal + 13, column=1).value = "JANTINA"
                        sheet.cell(row=row_klinik_asal + 14, column=1).value = "LELAKI"
                        sheet.cell(row=row_klinik_asal + 15, column=1).value = "PEREMPUAN"
                        
                        sheet.cell(row=row_klinik_asal + 18, column=1).value = "BANGSA"
                        sheet.cell(row=row_klinik_asal + 19, column=1).value = "MELAYU"
                        sheet.cell(row=row_klinik_asal + 20, column=1).value = "CINA"
                        sheet.cell(row=row_klinik_asal + 21, column=1).value = "INDIA"
                        sheet.cell(row=row_klinik_asal + 22, column=1).value = "BUMIPUTERA"
                        sheet.cell(row=row_klinik_asal + 23, column=1).value = "WARGA ASING"
                        
                        sheet.cell(row=row_klinik_asal + 26, column=1).value = "BANGSA"
                        sheet.cell(row=row_klinik_asal + 27, column=1).value = "CD [1]"
                        sheet.cell(row=row_klinik_asal + 28, column=1).value = "CD [2]"
                        sheet.cell(row=row_klinik_asal + 29, column=1).value = "FILEM 14X17 [1]"
                        sheet.cell(row=row_klinik_asal + 30, column=1).value = "FILEM 14X17 [2]"
                        sheet.cell(row=row_klinik_asal + 31, column=1).value = "FILEM 10X12 [1]"
                        sheet.cell(row=row_klinik_asal + 32, column=1).value = "FILEM 10X12 [2]"
                        
                        # Kemas kini semua formula rujukan harian (Kolum B hingga AF / 2 hingga 32)
                        for d in range(1, 32):
                            col_idx = d + 1
                            # Klinik Rujukan Luar
                            for idx in range(ref_rows_count):
                                r_row = 14 + idx
                                sheet.cell(row=r_row, column=col_idx).value = f"=SUM('{d}'!AA{r_row})"
                            
                            # Klinik Asal (Pesakit)
                            row_jumlah_asal_in_daily = 15 + ref_rows_count
                            sheet.cell(row=row_klinik_asal, column=col_idx).value = f"=SUM('{d}'!AA{row_jumlah_asal_in_daily})"
                            
                            # Jumlah Pesakit
                            col_letter = openpyxl.utils.get_column_letter(col_idx)
                            sheet.cell(row=row_klinik_asal + 1, column=col_idx).value = f"=SUM({col_letter}14:{col_letter}{row_klinik_asal})"
                            
                            # TB singkatan_klinik & TB Klinik Luar
                            sheet.cell(row=row_klinik_asal + 4, column=col_idx).value = f"=SUM('{d}'!V38)"
                            sheet.cell(row=row_klinik_asal + 5, column=col_idx).value = f"=SUM('{d}'!V39)"
                            
                            # Wheelchair / Trolley
                            sheet.cell(row=row_klinik_asal + 9, column=col_idx).value = f"=SUM('{d}'!V34)"
                            sheet.cell(row=row_klinik_asal + 10, column=col_idx).value = f"=SUM('{d}'!V35)"
                            
                            # Jantina (Lelaki / Perempuan)
                            sheet.cell(row=row_klinik_asal + 14, column=col_idx).value = f"=SUM('{d}'!V8)"
                            sheet.cell(row=row_klinik_asal + 15, column=col_idx).value = f"=SUM('{d}'!V9)"
                            
                            # Bangsa
                            sheet.cell(row=row_klinik_asal + 19, column=col_idx).value = f"=SUM('{d}'!Y6)"
                            sheet.cell(row=row_klinik_asal + 20, column=col_idx).value = f"=SUM('{d}'!Y7)"
                            sheet.cell(row=row_klinik_asal + 21, column=col_idx).value = f"=SUM('{d}'!Y8)"
                            sheet.cell(row=row_klinik_asal + 22, column=col_idx).value = f"=SUM('{d}'!Y9)"
                            sheet.cell(row=row_klinik_asal + 23, column=col_idx).value = f"=SUM('{d}'!Y10)"
                            
                            # Media / Consumables
                            sheet.cell(row=row_klinik_asal + 27, column=col_idx).value = f"=SUM('{d}'!AA28)"
                            sheet.cell(row=row_klinik_asal + 28, column=col_idx).value = f"=SUM('{d}'!AA31*2)"
                            sheet.cell(row=row_klinik_asal + 29, column=col_idx).value = f"=SUM('{d}'!AA29)"
                            sheet.cell(row=row_klinik_asal + 30, column=col_idx).value = f"=SUM('{d}'!AA32*2)"
                            sheet.cell(row=row_klinik_asal + 31, column=col_idx).value = f"=SUM('{d}'!AA30)"
                            sheet.cell(row=row_klinik_asal + 32, column=col_idx).value = f"=SUM('{d}'!AA33*2)"
                        
                    # C. Kemas kini BIL PT (D2 & H2)
                    if "BIL PT" in wb.sheetnames:
                        sheet = wb["BIL PT"]
                        sheet.cell(row=2, column=4).value = singkatan_klinik
                        sheet.cell(row=2, column=8).value = singkatan_klinik
                        
                    # D. Kemas kini Reten Harian (A3)
                    if "Reten Harian" in wb.sheetnames:
                        sheet = wb["Reten Harian"]
                        sheet.cell(row=3, column=1).value = singkatan_klinik
                        
                    # E. Kemas kini helaian harian ('1' hingga '31')
                    for sheetname in wb.sheetnames:
                        if sheetname.isdigit():
                            sheet = wb[sheetname]
                            
                            # 1. Tajuk di sel A1
                            sheet.cell(row=1, column=1).value = f"BUKU DAFTAR RADIOLOGI {singkatan_klinik}"
                            
                            # 2. Singkatan klinik & formula di Kolum U & V
                            sheet.cell(row=25, column=21).value = singkatan_klinik  # U25
                            sheet.cell(row=30, column=21).value = singkatan_klinik  # U30
                            sheet.cell(row=38, column=21).value = singkatan_klinik  # U38
                            sheet.cell(row=13, column=22).value = singkatan_klinik  # V13
                            sheet.cell(row=14, column=22).value = f'=COUNTIFS(L9:L73, "DADA", M9:M73, "CXR", O9:O73, "{klinik_asal}")' # V14
                            sheet.cell(row=25, column=22).value = f'=COUNTIF(O9:O73, "{klinik_asal}")' # V25
                            
                            # 3. Kosongkan senarai rujukan lama di Kolum Z & AA (14 hingga 50)
                            for r in range(14, 50):
                                sheet.cell(row=r, column=26).value = None
                                sheet.cell(row=r, column=27).value = None
                                
                            # 4. Tulis senarai klinik luar (Kolum Z & AA)
                            for idx in range(ref_rows_count):
                                r_row = 14 + idx
                                if idx < num_ref:
                                    c_name = referrals_only[idx]
                                    sheet.cell(row=r_row, column=26).value = c_name
                                    sheet.cell(row=r_row, column=27).value = f'=COUNTIF(O9:O73, "{c_name}")'
                                else:
                                    sheet.cell(row=r_row, column=26).value = ""
                                    sheet.cell(row=r_row, column=27).value = 0
                                    
                            # 5. Jumlah klinik luar & Jumlah klinik asal
                            row_jumlah_luar = 14 + ref_rows_count
                            sheet.cell(row=row_jumlah_luar, column=26).value = "JUMLAH KLINIK LUAR"
                            sheet.cell(row=row_jumlah_luar, column=27).value = f'=SUM(AA14:AA{13+ref_rows_count})'
                            
                            row_jumlah_asal = row_jumlah_luar + 1
                            sheet.cell(row=row_jumlah_asal, column=26).value = f"JUMLAH {singkatan_klinik}"
                            sheet.cell(row=row_jumlah_asal, column=27).value = f'=COUNTIF(O9:O73, "{klinik_asal}")'
                            
                    wb.save(filepath)
                    wb.close()
                    del wb
                    updated_files += 1
                except Exception as e:
                    print(f"Ralat mengemas kini data di dalam {filepath}: {e}")
                    
    gc.collect()
    print(f"Selesai menyelaraskan Excel. {updated_files} fail dikemas kini.")
    return True

# Utiliti: Membaca pilihan dropdown daripada Masterlist Excel
def get_masterlist_options():
    config = load_config()
    
    options = {
        "bahagian": [
            "CXR", "CXR RME", "CXR TB", "AXR", "KUB", "ANKLE", "CALCANEUM", "CLAVICLE", 
            "ELBOW", "FEMUR", "FINGER", "FOOT", "HAND", "HIP", "HUMERUS", "KNEE", 
            "MORTISE", "PATELLA", "RADIUS ULNA", "SCAPHOID", "SCAPULA Y", "SHOULDER", 
            "SKYLINE", "THUMB", "TIBIA FIBULA", "TOE", "WRIST", "FACE (OMV/SMV/LAT)", 
            "MANDIBLE", "NASAL", "ORBIT", "PNS", "SKULL (AP/LAT)", "TMJ", "CERVICAL (AP/LAT)", 
            "COCCYX", "LUMBOSACRAL", "NECK", "SACROILIAC JT", "SACRUM", "THORACIC", 
            "THORACOLUMBAR", "WHOLESPINE", "PELVIS", "HIP", "LAIN-LAIN"
        ],
        "klinik": config.get("klinik_rujukan", ["KK SELAYANG BARU", "KK GOMBAK SETIA", "KK TAMAN EHSAN", "LAIN-LAIN"]),
        "warganegara": ["YA", "TIDAK"],
        "jenis": [
            "DADA", "ABDOMEN", "EXTREMITI", "RANGKA KEPALA", "SPINA VERTEBRA", "PELVIS", 
            "DADA REF", "ABDOMEN REF", "EXTREMITI REF", "RANGKA KEPALA REF", "SPINA VERTEBRA REF", 
            "PELVIS REF", "CXR RME KK LUAR", "CXR TB KK LUAR"
        ],
        "kategori": ["WALK-IN", "WHEELCHAIR", "TROLLEY"],
        "kanankiri": ["KANAN", "KIRI", "KEDUA-DUA", "-"],
        "kakitangan": ["YA", "TIDAK"],
        "kaum": ["MELAYU", "CINA", "INDIA", "BUMIPUTERA", "WARGA ASING", "LAIN-LAIN"],
        "cd_filem": ["CD", "FILEM 14X17", "FILEM 10X12", "TIADA"]
    }
    
    # Gunakan template Januari tahun semasa untuk membaca Masterlist tambahan jika ada
    current_year = datetime.date.today().year
    temp_date = datetime.date(current_year, 1, 1)
    filepath = get_excel_path(temp_date)
    
    # Cuba tahun lain jika tahun semasa tidak ditemui
    if not filepath or not os.path.exists(filepath):
        if os.path.exists(PENDAFTARAN_DIR):
            try:
                years = [d for d in os.listdir(PENDAFTARAN_DIR) if d.isdigit()]
                if years:
                    temp_date = datetime.date(int(sorted(years)[0]), 1, 1)
                    filepath = get_excel_path(temp_date)
            except Exception:
                pass
                
    if not filepath or not os.path.exists(filepath):
        return options
        
    try:
        excel_options = {k: [] for k in options.keys() if k != "klinik"}
        excel_options["klinik"] = options["klinik"]
        
        wb = openpyxl.load_workbook(filepath, read_only=True)
        if "Masterlist" in wb.sheetnames:
            sheet = wb["Masterlist"]
            
            for r_idx in range(4, 50):
                val_b = sheet.cell(row=r_idx, column=2).value
                if val_b: excel_options["bahagian"].append(str(val_b).strip())
                
                val_d = sheet.cell(row=r_idx, column=4).value
                if val_d: excel_options["warganegara"].append(str(val_d).strip())
                
                val_f = sheet.cell(row=r_idx, column=6).value
                if val_f: excel_options["kategori"].append(str(val_f).strip())
                
                val_h = sheet.cell(row=r_idx, column=8).value
                if val_h: excel_options["kanankiri"].append(str(val_h).strip())
                
                val_j = sheet.cell(row=r_idx, column=10).value
                if val_j: excel_options["kakitangan"].append(str(val_j).strip())
                
                val_m = sheet.cell(row=r_idx, column=13).value
                if val_m: excel_options["kaum"].append(str(val_m).strip())
                
                val_o = sheet.cell(row=r_idx, column=15).value
                if val_o: excel_options["cd_filem"].append(str(val_o).strip())

            for r_idx in range(9, 25):
                val_e = sheet.cell(row=r_idx, column=5).value
                if val_e: excel_options["jenis"].append(str(val_e).strip())
                
            # Guna data dari Excel jika ada kandungan diisi
            for k in excel_options:
                if excel_options[k] and k != "klinik":
                    options[k] = list(dict.fromkeys(excel_options[k]))
        wb.close()
    except Exception as e:
        print(f"Error reading masterlist: {e}")
        
    return options

# Utiliti: Mengira Nombor X-ray seterusnya secara auto-increment ke belakang
def get_next_xray_no(target_date):
    current_date = target_date
    # Had carian: Jangan cari ke belakang melebihi 1 tahun
    limit_date = target_date - datetime.timedelta(days=365)
    
    while current_date > limit_date:
        filepath = get_excel_path(current_date)
        if not filepath or not os.path.exists(filepath):
            # Cuba bulan sebelumnya
            # Dapatkan hari terakhir bulan sebelumnya
            first_day_of_month = datetime.date(current_date.year, current_date.month, 1)
            current_date = first_day_of_month - datetime.timedelta(days=1)
            continue
            
        try:
            # Membuka fail secara read_only untuk prestasi cepat
            wb = openpyxl.load_workbook(filepath, read_only=True)
            # Imbas helaian (sheets) bermula dari hari semasa ke belakang (hanya helaian berbentuk nombor hari)
            day_sheets = [s for s in wb.sheetnames if s.isdigit()]
            # Susun mengikut nombor hari menurun (contoh: 31, 30, ... 1)
            day_sheets = sorted(day_sheets, key=int, reverse=True)
            
            # Jika fail yang dibuka adalah bulan yang sama dengan target_date,
            # hanya semak helaian dari target_date.day ke bawah
            if current_date.year == target_date.year and current_date.month == target_date.month:
                day_sheets = [s for s in day_sheets if int(s) <= target_date.day]
                
            for sheetname in day_sheets:
                sheet = wb[sheetname]
                # Imbas dari baris 73 ke 9 ke belakang secara optimum menggunakan iter_rows
                rows_data = list(sheet.iter_rows(min_row=9, max_row=73, min_col=3, max_col=3, values_only=True))
                for val_tuple in reversed(rows_data):
                    xray_val = val_tuple[0]
                    if xray_val is not None:
                        # Buang sebarang aksara bukan digit untuk increment
                        try:
                            clean_val = "".join(filter(str.isdigit, str(xray_val)))
                            if clean_val:
                                wb.close()
                                return int(clean_val) + 1
                        except ValueError:
                            pass
            wb.close()
        except Exception as e:
            print(f"Error scanning file for X-ray auto-increment: {filepath}, error: {e}")
            
        # Pergi ke bulan sebelumnya jika bulan ini kosong
        first_day_of_month = datetime.date(current_date.year, current_date.month, 1)
        current_date = first_day_of_month - datetime.timedelta(days=1)
        
    return 1 # Default jika tiada rekod langsung ditemui

# 1. Halaman Borang Input
@app.route("/")
def index():
    config = load_config()
    if not config.get("is_configured", False):
        return redirect(url_for("setup"))
    today_str = datetime.date.today().isoformat()
    options = get_masterlist_options()
    return render_template("borang.html", today=today_str, options=options, klinik_asal=config.get("klinik_asal", "KK RAWANG PERDANA"))

# API: Dapatkan Nombor X-ray Seterusnya untuk Tarikh Tertentu
@app.route("/api/next-xray")
def api_next_xray():
    date_str = request.args.get("date")
    if not date_str:
        return jsonify({"error": "Tarikh diperlukan"}), 400
    try:
        date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
        next_no = get_next_xray_no(date_obj)
        # Format as 4-digit padded string (e.g., 0001, 0118, etc.)
        formatted_no = f"{next_no:04d}"
        return jsonify({"next_xray": formatted_no})
    except Exception as e:
        return jsonify({"error": str(e)}), 500

# API: Simpan Data Pendaftaran Pesakit
@app.route("/submit", methods=["POST"])
def submit_registration():
    config = load_config()
    try:
        # Mengambil input borang (sama ada JSON atau Form-data)
        if request.is_json:
            data = request.json
            tarikh_str = data.get("tarikh")
            id_type = data.get("id_type")
            id_number = data.get("id_number")
            name = data.get("name", "").upper()
            age = data.get("age")
            gender = data.get("gender")
            citizen = data.get("citizen")
            gov_staff = data.get("gov_staff")
            race = data.get("race", "").upper()
            address = data.get("address", "").upper()
            clinic = data.get("clinic", "").upper()
            category = data.get("category", "").upper()
            lmp = data.get("lmp") or "-"
            
            # Senarai pemeriksaan (untuk multi-part)
            exams = data.get("exams", [])
        else:
            tarikh_str = request.form.get("tarikh")
            id_type = request.form.get("id_type")
            id_number = request.form.get("id_number")
            name = request.form.get("name", "").upper()
            age = request.form.get("age")
            gender = request.form.get("gender")
            citizen = request.form.get("citizen")
            gov_staff = request.form.get("gov_staff")
            race = request.form.get("race", "").upper()
            address = request.form.get("address", "").upper()
            clinic = request.form.get("clinic", "").upper()
            category = request.form.get("category", "").upper()
            lmp = request.form.get("lmp") or "-"
            
            # Format tunggal dari form lama
            exams = [{
                "xray_no": request.form.get("xray_no"),
                "exam_type": request.form.get("exam_type", "").upper(),
                "exam_part": request.form.get("exam_part", "").upper(),
                "laterality": request.form.get("laterality", "").upper(),
                "cd_filem": request.form.get("cd_filem", "").upper(),
                "total_expose": request.form.get("total_expose") or 1
            }]
            
        # Formatkan tarikh LMP dari YYYY-MM-DD ke DD-MM-YYYY jika format kalendar digunakan
        if lmp and lmp != "-":
            try:
                lmp_date = datetime.datetime.strptime(lmp, "%Y-%m-%d").date()
                lmp = lmp_date.strftime("%d-%m-%Y")
            except ValueError:
                pass # Kekalkan jika format lain (cth: DD-MM-YYYY sedia ada)
        
        # Validasi tarikh
        date_obj = datetime.datetime.strptime(tarikh_str, "%Y-%m-%d").date()
        day = date_obj.day
        
        filepath = get_excel_path(date_obj)
        if not filepath or not os.path.exists(filepath):
            return jsonify({"success": False, "message": f"Fail Excel untuk tarikh {tarikh_str} tidak wujud!"}), 400
            
        # Membuka fail Excel untuk penulisan
        wb = openpyxl.load_workbook(filepath)
        sheetname = str(day)
        if sheetname not in wb.sheetnames:
            wb.close()
            return jsonify({"success": False, "message": f"Helaian (Sheet) untuk hari {day} tidak wujud dalam fail Excel!"}), 400
            
        sheet = wb[sheetname]
        
        # Semak ruang kosong yang mencukupi untuk semua pemeriksaan
        required_rows = len(exams)
        empty_rows = []
        for r_idx in range(9, 74):
            if sheet.cell(row=r_idx, column=3).value is None:
                empty_rows.append(r_idx)
                if len(empty_rows) == required_rows:
                    break
                    
        if len(empty_rows) < required_rows:
            wb.close()
            return jsonify({"success": False, "message": f"Ruang tidak mencukupi! Hanya tinggal {len(empty_rows)} baris kosong sahaja bagi hari {day}."}), 400
            
        # Tulis setiap pemeriksaan ke baris kosong berturutan
        is_mykad = (id_type == "MyKad" and len(id_number.replace("-", "")) == 12)
        
        for idx, exam in enumerate(exams):
            target_row = empty_rows[idx]
            
            # Col A: Tarikh
            sheet.cell(row=target_row, column=1).value = date_obj.strftime("%d-%m-%Y")
            # Col B: Bil Kes
            if sheet.cell(row=target_row, column=2).value is None:
                sheet.cell(row=target_row, column=2).value = target_row - 8
                
            # Ciri pemeriksaan khusus
            exam_xray_no = exam.get("xray_no")
            exam_type = exam.get("exam_type", "").upper()
            exam_part = exam.get("exam_part", "").upper()
            laterality = exam.get("laterality", "").upper()
            cd_filem = exam.get("cd_filem", "").upper()
            total_expose = exam.get("total_expose") or 1
            
            sheet.cell(row=target_row, column=3).value = exam_xray_no # Col C: Nombor X-ray
            sheet.cell(row=target_row, column=4).value = id_number # Col D: No KP / Pasport
            sheet.cell(row=target_row, column=5).value = name # Col E: Nama
            
            # Lajur F (UMUR) & G (JANTINA)
            if is_mykad:
                sheet.cell(row=target_row, column=6).value = f'=DATEDIF(DATE(LEFT(D{target_row},2),MONTH(MID(D{target_row},3,2)),DAY(MID(D{target_row},5,2))), NOW(), "Y")'
                sheet.cell(row=target_row, column=7).value = f'=IF(MOD(RIGHT(D{target_row}),2),"M","F")'
            else:
                sheet.cell(row=target_row, column=6).value = int(age) if age else ""
                sheet.cell(row=target_row, column=7).value = "M" if gender == "LELAKI" else "F"
                
            sheet.cell(row=target_row, column=8).value = citizen # Col H: W/NEGARA
            sheet.cell(row=target_row, column=9).value = gov_staff # Col I: K/TANGAN KERAJAAN
            sheet.cell(row=target_row, column=10).value = race # Col J: BANGSA
            sheet.cell(row=target_row, column=11).value = address # Col K: ALAMAT
            sheet.cell(row=target_row, column=12).value = exam_type # Col L: JENIS
            sheet.cell(row=target_row, column=13).value = exam_part # Col M: BAHAGIAN
            sheet.cell(row=target_row, column=14).value = laterality # Col N: LATERALLY
            sheet.cell(row=target_row, column=15).value = clinic # Col O: KLINIK
            sheet.cell(row=target_row, column=16).value = category # Col P: KATEGORI
            sheet.cell(row=target_row, column=17).value = lmp # Col Q: LMP
            sheet.cell(row=target_row, column=18).value = cd_filem # Col R: CD/FILEM
            
            # Semak jika klinik rujukan baru & kemas kini konfigurasi serta Excel secara automatik
            if clinic and clinic not in config.get("klinik_rujukan", []):
                config["klinik_rujukan"].append(clinic)
                save_config(config)
                threading.Thread(target=sync_excel_masterlists, args=(config["klinik_rujukan"], config.get("klinik_asal", ""), config.get("singkatan_klinik", "")), daemon=True).start()
            
            # Simpan dose (tanpa perpuluhan)
            try:
                sheet.cell(row=target_row, column=19).value = int(float(total_expose))
            except (ValueError, TypeError):
                sheet.cell(row=target_row, column=19).value = 1
                
        wb.save(filepath)
        wb.close()
        
        # Simpan Data Tambahan PHRIS (COVID, DM, Temujanji, Rejection) - Lalai 0
        extra_data = load_extra_phris_data()
        
        # Letakkan lalai 0 bagi data tambahan memandangkan panel PHRIS dibuang
        # melainkan ia dihantar (sokong keserasian)
        if request.is_json:
            covid_count = int(data.get("extra_covid") or 0)
            dm_count = int(data.get("extra_dm") or 0)
            usg_count = int(data.get("extra_usg") or 0)
            mammo_count = int(data.get("extra_mammo") or 0)
            ct_count = int(data.get("extra_ct") or 0)
            lain_count = int(data.get("extra_lain") or 0)
            
            r_details = data.get("penolakan", {})
            rejects = {
                "over_exposure": int(r_details.get("over_exposure") or 0),
                "under_exposure": int(r_details.get("under_exposure") or 0),
                "double_exposure": int(r_details.get("double_exposure") or 0),
                "wrong_technique": int(r_details.get("wrong_technique") or 0),
                "wrong_patient": int(r_details.get("wrong_patient") or 0),
                "wrong_marker": int(r_details.get("wrong_marker") or 0),
                "collimation_error": int(r_details.get("collimation_error") or 0),
                "patient_movement": int(r_details.get("patient_movement") or 0),
                "patient_artifact": int(r_details.get("patient_artifact") or 0),
                "equipment_fault": int(r_details.get("equipment_fault") or 0),
                "detector_fault": int(r_details.get("detector_fault") or 0),
                "image_artifact": int(r_details.get("image_artifact") or 0),
                "processing_fault": int(r_details.get("processing_fault") or 0),
                "miscellaneous": int(r_details.get("miscellaneous") or 0),
                "jumlah_imej": int(r_details.get("jumlah_imej") or 0),
                "pengulangan": int(r_details.get("pengulangan") or 0)
            }
        else:
            covid_count = int(request.form.get("extra_covid") or 0)
            dm_count = int(request.form.get("extra_dm") or 0)
            usg_count = int(request.form.get("extra_usg") or 0)
            mammo_count = int(request.form.get("extra_mammo") or 0)
            ct_count = int(request.form.get("extra_ct") or 0)
            lain_count = int(request.form.get("extra_lain") or 0)
            
            rejects = {
                "over_exposure": int(request.form.get("rej_over_exposure") or 0),
                "under_exposure": int(request.form.get("rej_under_exposure") or 0),
                "double_exposure": int(request.form.get("rej_double_exposure") or 0),
                "wrong_technique": int(request.form.get("rej_wrong_technique") or 0),
                "wrong_patient": int(request.form.get("rej_wrong_patient") or 0),
                "wrong_marker": int(request.form.get("rej_wrong_marker") or 0),
                "collimation_error": int(request.form.get("rej_collimation_error") or 0),
                "patient_movement": int(request.form.get("rej_patient_movement") or 0),
                "patient_artifact": int(request.form.get("rej_patient_artifact") or 0),
                "equipment_fault": int(request.form.get("rej_equipment_fault") or 0),
                "detector_fault": int(request.form.get("rej_detector_fault") or 0),
                "image_artifact": int(request.form.get("rej_image_artifact") or 0),
                "processing_fault": int(request.form.get("rej_processing_fault") or 0),
                "miscellaneous": int(request.form.get("rej_miscellaneous") or 0),
                "jumlah_imej": int(request.form.get("rej_jumlah_imej") or 0),
                "pengulangan": int(request.form.get("rej_pengulangan") or 0)
            }
            
        extra_data[tarikh_str] = {
            "covid": covid_count,
            "dm": dm_count,
            "temujanji": {
                "usg": usg_count,
                "mammo": mammo_count,
                "ct_scan": ct_count,
                "lain": lain_count
            },
            "penolakan": rejects
        }
        
        save_extra_phris_data(extra_data)
        
        return jsonify({"success": True, "message": "Pesakit telah didaftarkan!"})
        
    except Exception as e:
        return jsonify({"success": False, "message": f"Ralat semasa menyimpan: {str(e)}"}), 500

# 2. Halaman Dashboard
@app.route("/dashboard")
def dashboard():
    today = datetime.date.today()
    selected_month = request.args.get("month", default=today.month, type=int)
    selected_year = request.args.get("year", default=today.year, type=int)
    
    # Ambil senarai tahun yang tersedia dalam Pendaftaran
    available_years = []
    if os.path.exists(PENDAFTARAN_DIR):
        available_years = [int(d) for d in os.listdir(PENDAFTARAN_DIR) if d.isdigit()]
    if not available_years:
        available_years = [selected_year]
    available_years = sorted(available_years)
    
    return render_template("dashboard.html", 
                           selected_month=selected_month, 
                           selected_year=selected_year,
                           available_years=available_years,
                           months=MONTH_MAP)

# API: Dapatkan Data untuk Graf Dashboard
@app.route("/api/dashboard-data")
def api_dashboard_data():
    month = request.args.get("month", default=datetime.date.today().month, type=int)
    year = request.args.get("year", default=datetime.date.today().year, type=int)
    
    # Bina objek tarikh awal untuk bulan berkenaan
    temp_date = datetime.date(year, month, 1)
    filepath = get_excel_path(temp_date)
    
    # Objek data agregat
    data = {
      "daily_cases": [0] * 31,
      "exam_types": {},
      "exam_parts": {},
      "cases_genders": {"LELAKI": 0, "PEREMPUAN": 0},
      "cases_races": {},
      "cases_citizens": {"YA": 0, "TIDAK": 0},
      "patients_genders": {"LELAKI": 0, "PEREMPUAN": 0},
      "patients_races": {},
      "patients_citizens": {"YA": 0, "TIDAK": 0},
      "consumables": {"CD": 0, "FILEM 14X17": 0, "FILEM 10X12": 0},
      "clinics": {},
      "total_patients": 0,
      "total_cases": 0
    }
    
    if not filepath or not os.path.exists(filepath):
        return jsonify(data)
        
    try:
        # Guna data_only=True untuk membaca nilai sebenar daripada formula Excel
        wb = openpyxl.load_workbook(filepath, data_only=True)
        unique_patients = {}
        
        for sheetname in wb.sheetnames:
            if sheetname.isdigit():
                day_idx = int(sheetname) - 1
                sheet = wb[sheetname]
                
                # Baca baris 9 hingga 73
                for r_idx in range(9, 74):
                    xray_no = sheet.cell(row=r_idx, column=3).value
                    if xray_no is not None:
                        data["daily_cases"][day_idx] += 1
                        data["total_cases"] += 1
                        
                        # Dapatkan ID dan Nama untuk penentu unik pesakit
                        id_number = sheet.cell(row=r_idx, column=4).value or ""
                        name = sheet.cell(row=r_idx, column=5).value or ""
                        patient_key = (str(id_number).strip(), str(name).strip().upper())
                        
                        # Jantina (Col G)
                        gender_val = sheet.cell(row=r_idx, column=7).value
                        g_str = ""
                        if gender_val:
                            g_str = str(gender_val).strip().upper()
                        if not g_str or g_str.startswith("="):
                            parsed = parse_mykad_py(str(id_number))
                            if parsed:
                                g_str = "L" if parsed["is_male"] else "P"
                        
                        gender_clean = "LELAKI" if g_str in ["M", "L", "LELAKI", "MALE"] else ("PEREMPUAN" if g_str in ["F", "P", "PEREMPUAN", "FEMALE"] else "TIADA")
                        
                        # Bangsa (Col J)
                        bangsa_val = sheet.cell(row=r_idx, column=10).value
                        race_clean = str(bangsa_val).strip().upper() if bangsa_val else "TIADA"
                        
                        # Warganegara (Col H)
                        citizen_val = sheet.cell(row=r_idx, column=8).value
                        citizen_clean = str(citizen_val).strip().upper() if citizen_val else "YA"
                        if citizen_clean not in ["YA", "TIDAK"]:
                            citizen_clean = "YA"
                        
                        # Increment case counters
                        if gender_clean != "TIADA":
                            data["cases_genders"][gender_clean] += 1
                        if race_clean != "TIADA":
                            data["cases_races"][race_clean] = data["cases_races"].get(race_clean, 0) + 1
                        data["cases_citizens"][citizen_clean] += 1
                        
                        # Store in unique patients dictionary
                        if patient_key not in unique_patients:
                            unique_patients[patient_key] = {
                                "gender": gender_clean,
                                "race": race_clean,
                                "citizen": citizen_clean
                            }
                        
                        # Jenis Pemeriksaan (Col L)
                        j_val = sheet.cell(row=r_idx, column=12).value
                        if j_val:
                            j_val_str = str(j_val).strip().upper()
                            # Bersihkan ref jika ada untuk dikelompokkan
                            j_clean = j_val_str.replace(" REF", "").replace(" KK LUAR", "")
                            data["exam_types"][j_clean] = data["exam_types"].get(j_clean, 0) + 1
                            
                        # Bahagian Pemeriksaan (Col M)
                        part_val = sheet.cell(row=r_idx, column=13).value
                        if part_val:
                            part_str = str(part_val).strip().upper()
                            data["exam_parts"][part_str] = data["exam_parts"].get(part_str, 0) + 1
                            
                        # Klinik Rujukan (Col O)
                        klinik_val = sheet.cell(row=r_idx, column=15).value
                        if klinik_val:
                            klinik_name = str(klinik_val).strip().upper()
                            data["clinics"][klinik_name] = data["clinics"].get(klinik_name, 0) + 1
                        else:
                            data["clinics"]["TIADA"] = data["clinics"].get("TIADA", 0) + 1
                                
                        # Consumable (Col R: CD/FILEM)
                        cons_val = sheet.cell(row=r_idx, column=18).value
                        if cons_val:
                            cons_str = str(cons_val).strip().upper()
                            if "CD" in cons_str:
                                # CD [1] atau CD [2]
                                count = 2 if "[2]" in cons_str else 1
                                data["consumables"]["CD"] += count
                            elif "14X17" in cons_str:
                                count = 2 if "[2]" in cons_str else 1
                                data["consumables"]["FILEM 14X17"] += count
                            elif "10X12" in cons_str:
                                count = 2 if "[2]" in cons_str else 1
                                data["consumables"]["FILEM 10X12"] += count
                                
        # Populate patient counters from unique_patients dictionary
        for p_info in unique_patients.values():
            g = p_info["gender"]
            r = p_info["race"]
            c = p_info["citizen"]
            
            if g != "TIADA":
                data["patients_genders"][g] += 1
            if r != "TIADA":
                data["patients_races"][r] = data["patients_races"].get(r, 0) + 1
            data["patients_citizens"][c] += 1
            
        data["total_patients"] = len(unique_patients)
        wb.close()
    except Exception as e:
        print(f"Error generating dashboard data: {e}")
        
    return jsonify(data)

# 3. Halaman Preview Reten PHRIS
@app.route("/phris")
def phris():
    today = datetime.date.today()
    selected_year = request.args.get("year", default=today.year, type=int)
    
    # Ambil senarai tahun yang tersedia dalam Pendaftaran
    available_years = []
    if os.path.exists(PENDAFTARAN_DIR):
        available_years = [int(d) for d in os.listdir(PENDAFTARAN_DIR) if d.isdigit()]
    if not available_years:
        available_years = [selected_year]
    available_years = sorted(available_years)
    
    return render_template("reten_phris.html", 
                           selected_year=selected_year,
                           available_years=available_years)

# API: Menjana Matriks Data PHRIS (JAN-DIS) bagi Tahun Tertentu
@app.route("/api/phris-matrix")
def api_phris_matrix():
    year = request.args.get("year", default=datetime.date.today().year, type=int)
    extra_data = load_extra_phris_data()
    config = load_config()
    klinik_asal = config.get("klinik_asal", "KK RAWANG PERDANA").upper()
    
    # Struktur Data Reten Matriks (JAN - DIS)
    # Setiap kategori mempunyai senarai 12 elemen (mewakili bulan 1 hingga 12)
    matrix = {
        "bangsa": {
            "MELAYU": [0] * 12,
            "CINA": [0] * 12,
            "INDIA": [0] * 12,
            "BUMIPUTERA": [0] * 12,
            "WARGA ASING": [0] * 12,
            "LAIN-LAIN": [0] * 12,
        },
        "kedatangan": {
            "TROLLEY": [0] * 12,
            "WHEELCHAIR": [0] * 12,
            "RUJUK TERUS": [0] * 12,
            "KLINIK_OPD": [0] * 12, # Walk-in pesakit
        },
        "pemeriksaan_am": {
            "DADA": [0] * 12,
            "ABDOMEN": [0] * 12,
            "EXTREMITI": [0] * 12,
            "RANGKA KEPALA": [0] * 12,
            "SPINA VERTEBRA": [0] * 12,
            "PELVIS": [0] * 12,
            "SKELETAL SURVEY": [0] * 12,
            "DEXA": [0] * 12,
            "OPG": [0] * 12,
            "LAIN-LAIN": [0] * 12,
        },
        "pemeriksaan_lain": {
            "RME": [0] * 12,
            "PTB": [0] * 12,
            "KES LAIN": [0] * 12,
            "COVID": [0] * 12, # Dari JSON
            "DM": [0] * 12,    # Dari JSON
        },
        "temujanji": {
            "USG": [0] * 12,     # Dari JSON
            "MAMMO": [0] * 12,   # Dari JSON
            "LAIN-LAIN": [0] * 12, # Dari JSON
            "CT-SCAN": [0] * 12, # Dari JSON
        },
        "consumable": {
            "CD-R": [0] * 12,
            "FILEM 10X12": [0] * 12,
            "FILEM 14X17": [0] * 12,
        },
        "penolakan": {
            "OVER EXPOSURE": [0] * 12,
            "UNDER EXPOSURE": [0] * 12,
            "DOUBLE EXPOSURE": [0] * 12,
            "WRONG TECHNIQUE": [0] * 12,
            "WRONG PATIENT": [0] * 12,
            "WRONG MARKER": [0] * 12,
            "COLLIMATION ERROR": [0] * 12,
            "PATIENT MOVEMENT": [0] * 12,
            "PATIENT ARTIFACT": [0] * 12,
            "EQUIPMENT FAULT": [0] * 12,
            "DETECTOR FAULT": [0] * 12,
            "IMAGE ARTIFACT": [0] * 12,
            "PROCESSING FAULT": [0] * 12,
            "MISCELLANEOUS": [0] * 12,
            "JUMLAH IMEJ": [0] * 12,
            "PENGULANGAN": [0] * 12,
        }
    }
    
    # Imbas semua fail 12 bulan dalam Excel
    for m in range(1, 13):
        m_idx = m - 1
        temp_date = datetime.date(year, m, 1)
        filepath = get_excel_path(temp_date)
        
        if not filepath or not os.path.exists(filepath):
            continue
            
        try:
            wb = openpyxl.load_workbook(filepath, data_only=True)
            for sheetname in wb.sheetnames:
                if sheetname.isdigit():
                    sheet = wb[sheetname]
                    for r_idx in range(9, 74):
                        xray_no = sheet.cell(row=r_idx, column=3).value
                        if xray_no is not None:
                            # 1. Agregasi Bangsa (Col J)
                            b_val = sheet.cell(row=r_idx, column=10).value
                            if b_val:
                                b_str = str(b_val).strip().upper()
                                if b_str in matrix["bangsa"]:
                                    matrix["bangsa"][b_str][m_idx] += 1
                                else:
                                    matrix["bangsa"]["LAIN-LAIN"][m_idx] += 1
                            else:
                                matrix["bangsa"]["LAIN-LAIN"][m_idx] += 1
                                
                            # 2. Agregasi Kedatangan (Col P & Col O)
                            kat_val = sheet.cell(row=r_idx, column=16).value
                            klinik_val = sheet.cell(row=r_idx, column=15).value
                            
                            # Kedatangan Kategori
                            kat_str = str(kat_val).strip().upper() if kat_val else ""
                            klinik_str = str(klinik_val).strip().upper() if klinik_val else ""
                            
                            # Tentukan jika Rujuk Terus (Klinik luar / Bukan klinik sendiri)
                            is_referral = (klinik_str != "" and klinik_str != klinik_asal)
                            
                            if kat_str == "TROLLEY":
                                matrix["kedatangan"]["TROLLEY"][m_idx] += 1
                            elif kat_str == "WHEELCHAIR":
                                matrix["kedatangan"]["WHEELCHAIR"][m_idx] += 1
                            
                            if is_referral:
                                matrix["kedatangan"]["RUJUK TERUS"][m_idx] += 1
                            else:
                                matrix["kedatangan"]["KLINIK_OPD"][m_idx] += 1
                                
                            # 3. Pemeriksaan AM (Col L)
                            jenis_val = sheet.cell(row=r_idx, column=12).value
                            if jenis_val:
                                j_str = str(jenis_val).strip().upper()
                                # Bersihkan rujukan
                                j_clean = j_str.replace(" REF", "").replace(" KK LUAR", "")
                                if j_clean in matrix["pemeriksaan_am"]:
                                    matrix["pemeriksaan_am"][j_clean][m_idx] += 1
                                else:
                                    matrix["pemeriksaan_am"]["LAIN-LAIN"][m_idx] += 1
                            else:
                                matrix["pemeriksaan_am"]["LAIN-LAIN"][m_idx] += 1
                                
                            # 4. Pemeriksaan Lain (Col M - BAHAGIAN & Col L)
                            bahagian_val = sheet.cell(row=r_idx, column=13).value
                            bah_str = str(bahagian_val).strip().upper() if bahagian_val else ""
                            
                            is_rme = ("RME" in bah_str or "RME" in j_str)
                            is_ptb = ("TB" in bah_str or "TB" in j_str)
                            
                            if is_rme:
                                matrix["pemeriksaan_lain"]["RME"][m_idx] += 1
                            elif is_ptb:
                                matrix["pemeriksaan_lain"]["PTB"][m_idx] += 1
                            else:
                                matrix["pemeriksaan_lain"]["KES LAIN"][m_idx] += 1
                                
                            # 5. Consumables (Col R)
                            cons_val = sheet.cell(row=r_idx, column=18).value
                            if cons_val:
                                cons_str = str(cons_val).strip().upper()
                                count = 2 if "[2]" in cons_str else (1 if "[1]" in cons_str or "[0]" not in cons_str else 0)
                                if "CD" in cons_str:
                                    matrix["consumable"]["CD-R"][m_idx] += count
                                elif "10X12" in cons_str:
                                    matrix["consumable"]["FILEM 10X12"][m_idx] += count
                                elif "14X17" in cons_str:
                                    matrix["consumable"]["FILEM 14X17"][m_idx] += count
            wb.close()
        except Exception as e:
            print(f"Error loading month {m} for PHRIS: {e}")
            
    # Gabung data tambahan PHRIS (COVID, DM, Temujanji, Rejections) dari JSON Sidecar
    for date_str, details in extra_data.items():
        try:
            # Semak jika tahun sepadan
            date_obj = datetime.datetime.strptime(date_str, "%Y-%m-%d").date()
            if date_obj.year != year:
                continue
                
            m_idx = date_obj.month - 1
            
            # Pemeriksaan Lain
            matrix["pemeriksaan_lain"]["COVID"][m_idx] += details.get("covid", 0)
            matrix["pemeriksaan_lain"]["DM"][m_idx] += details.get("dm", 0)
            
            # Temujanji
            t_details = details.get("temujanji", {})
            matrix["temujanji"]["USG"][m_idx] += t_details.get("usg", 0)
            matrix["temujanji"]["MAMMO"][m_idx] += t_details.get("mammo", 0)
            matrix["temujanji"]["CT-SCAN"][m_idx] += t_details.get("ct_scan", 0)
            matrix["temujanji"]["LAIN-LAIN"][m_idx] += t_details.get("lain", 0)
            
            # Rejection Analysis
            r_details = details.get("penolakan", {})
            matrix["penolakan"]["OVER EXPOSURE"][m_idx] += r_details.get("over_exposure", 0)
            matrix["penolakan"]["UNDER EXPOSURE"][m_idx] += r_details.get("under_exposure", 0)
            matrix["penolakan"]["DOUBLE EXPOSURE"][m_idx] += r_details.get("double_exposure", 0)
            matrix["penolakan"]["WRONG TECHNIQUE"][m_idx] += r_details.get("wrong_technique", 0)
            matrix["penolakan"]["WRONG PATIENT"][m_idx] += r_details.get("wrong_patient", 0)
            matrix["penolakan"]["WRONG MARKER"][m_idx] += r_details.get("wrong_marker", 0)
            matrix["penolakan"]["COLLIMATION ERROR"][m_idx] += r_details.get("collimation_error", 0)
            matrix["penolakan"]["PATIENT MOVEMENT"][m_idx] += r_details.get("patient_movement", 0)
            matrix["penolakan"]["PATIENT ARTIFACT"][m_idx] += r_details.get("patient_artifact", 0)
            matrix["penolakan"]["EQUIPMENT FAULT"][m_idx] += r_details.get("equipment_fault", 0)
            matrix["penolakan"]["DETECTOR FAULT"][m_idx] += r_details.get("detector_fault", 0)
            matrix["penolakan"]["IMAGE ARTIFACT"][m_idx] += r_details.get("image_artifact", 0)
            matrix["penolakan"]["PROCESSING FAULT"][m_idx] += r_details.get("processing_fault", 0)
            matrix["penolakan"]["MISCELLANEOUS"][m_idx] += r_details.get("miscellaneous", 0)
            matrix["penolakan"]["JUMLAH IMEJ"][m_idx] += r_details.get("jumlah_imej", 0)
            matrix["penolakan"]["PENGULANGAN"][m_idx] += r_details.get("pengulangan", 0)
            
        except Exception as e:
            print(f"Error parsing JSON date {date_str} for PHRIS: {e}")
            
    return jsonify(matrix)

# Utiliti: Jana fail Excel bagi suatu tahun jika belum wujud
def generate_excel_files_for_year(year, klinik_rujukan, klinik_asal, singkatan_klinik):
    referrals_only = [c for c in klinik_rujukan if c != klinik_asal]
    for month in range(1, 12 + 1):
        folder_name, file_month = MONTH_MAP[month]
        dir_path = os.path.join(PENDAFTARAN_DIR, str(year), folder_name)
        os.makedirs(dir_path, exist_ok=True)
        
        # Semak jika ada fail Excel di dalam folder
        files_exist = False
        try:
            for f in os.listdir(dir_path):
                if f.endswith(".xlsx") and not f.startswith("~$"):
                    files_exist = True
                    break
        except Exception:
            pass
            
        if not files_exist:
            dest_filename = f"{year} {file_month} PER.SS-RA 101 {singkatan_klinik}.xlsx"
            dest_path = os.path.join(dir_path, dest_filename)
            try:
                shutil.copy(TEMPLATE_FILE_PATH, dest_path)
                print(f"Jana fail database: {dest_path}")
            except Exception as e:
                print(f"Gagal jana fail database bulanan ke {dest_path}: {e}")

# 4. Halaman Setup Pertama Kali
@app.route("/setup")
def setup():
    config = load_config()
    return render_template("setup.html", config=config)

@app.route("/api/setup", methods=["POST"])
def api_setup():
    try:
        data = request.json
        klinik_asal = data.get("klinik_asal", "").strip().upper()
        singkatan_klinik = data.get("singkatan_klinik", "").strip().upper()
        klinik_rujukan = [c.strip().upper() for c in data.get("klinik_rujukan", []) if c.strip()]
        start_year = int(data.get("start_year", 2026))
        end_year = int(data.get("end_year", 2027))
        
        if not klinik_asal:
            return jsonify({"success": False, "message": "Nama Klinik Asal diperlukan!"}), 400
        if not singkatan_klinik:
            return jsonify({"success": False, "message": "Singkatan Klinik diperlukan!"}), 400
            
        if klinik_asal not in klinik_rujukan:
            klinik_rujukan.insert(0, klinik_asal)
            
        # Jana fail database excel dahulu bagi julat tahun yang dipilih
        for y in range(start_year, end_year + 1):
            generate_excel_files_for_year(y, klinik_rujukan, klinik_asal, singkatan_klinik)
            
        config_data = {
            "is_configured": True,
            "klinik_asal": klinik_asal,
            "singkatan_klinik": singkatan_klinik,
            "klinik_rujukan": klinik_rujukan
        }
        
        if save_config(config_data):
            threading.Thread(target=sync_excel_masterlists, args=(klinik_rujukan, klinik_asal, singkatan_klinik), daemon=True).start()
            return jsonify({"success": True, "message": "Konfigurasi berjaya disimpan!"})
        else:
            return jsonify({"success": False, "message": "Gagal menyimpan fail konfigurasi!"}), 500
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

# API: Tambah Klinik Rujukan Baru on-the-fly
@app.route("/api/settings/add-clinic", methods=["POST"])
def api_add_clinic():
    try:
        data = request.json
        new_clinic = data.get("clinic_name", "").strip().upper()
        if not new_clinic:
            return jsonify({"success": False, "message": "Nama klinik tidak boleh kosong!"}), 400
            
        config = load_config()
        if new_clinic in config.get("klinik_rujukan", []):
            return jsonify({"success": False, "message": "Klinik ini sudah wujud dalam senarai!"}), 400
            
        config["klinik_rujukan"].append(new_clinic)
        if save_config(config):
            threading.Thread(target=sync_excel_masterlists, args=(config["klinik_rujukan"], config["klinik_asal"], config.get("singkatan_klinik", "KKRP")), daemon=True).start()
            return jsonify({"success": True, "message": f"Klinik {new_clinic} berjaya ditambah!", "clinics": config["klinik_rujukan"]})
        else:
            return jsonify({"success": False, "message": "Gagal menyimpan konfigurasi!"}), 500
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

# Halaman Tetapan (Settings)
@app.route("/settings")
def settings_page():
    config = load_config()
    if not config.get("is_configured", False):
        return redirect(url_for("setup"))
    return render_template("settings.html", config=config)

# API: Simpan Tetapan Konfigurasi
@app.route("/api/settings/save", methods=["POST"])
def api_settings_save():
    try:
        data = request.json
        klinik_asal = data.get("klinik_asal", "").strip().upper()
        singkatan_klinik = data.get("singkatan_klinik", "").strip().upper()
        klinik_rujukan = [c.strip().upper() for c in data.get("klinik_rujukan", []) if c.strip()]
        
        if not klinik_asal or not singkatan_klinik:
            return jsonify({"success": False, "message": "Konfigurasi tidak lengkap!"}), 400
            
        config_data = {
            "is_configured": True,
            "klinik_asal": klinik_asal,
            "singkatan_klinik": singkatan_klinik,
            "klinik_rujukan": klinik_rujukan
        }
        
        if save_config(config_data):
            threading.Thread(target=sync_excel_masterlists, args=(klinik_rujukan, klinik_asal, singkatan_klinik), daemon=True).start()
            return jsonify({"success": True, "message": "Konfigurasi berjaya dikemaskini!"})
        else:
            return jsonify({"success": False, "message": "Gagal menyimpan konfigurasi!"}), 500
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

# API: Jana fail Excel bagi tahun baru (Database upgrade)
@app.route("/api/settings/generate-year", methods=["POST"])
def api_settings_generate_year():
    try:
        data = request.json
        year = int(data.get("year", 0))
        if not year:
            return jsonify({"success": False, "message": "Tahun tidak sah!"}), 400
            
        config = load_config()
        klinik_asal = config.get("klinik_asal", "")
        singkatan_klinik = config.get("singkatan_klinik", "")
        klinik_rujukan = config.get("klinik_rujukan", [])
        
        if not klinik_asal:
            return jsonify({"success": False, "message": "Konfigurasi sistem belum disetup!"}), 400
            
        generate_excel_files_for_year(year, klinik_rujukan, klinik_asal, singkatan_klinik)
        # Jalankan sync untuk fail baru
        threading.Thread(target=sync_excel_masterlists, args=(klinik_rujukan, klinik_asal, singkatan_klinik), daemon=True).start()
        
        return jsonify({"success": True, "message": f"Database untuk tahun {year} berjaya dijana!"})
    except Exception as e:
        return jsonify({"success": False, "message": str(e)}), 500

# 5. Halaman Visualisasi Data Penuh & Perbandingan Rujukan
@app.route("/data-penuh")
def data_penuh():
    config = load_config()
    if not config.get("is_configured", False):
        return redirect(url_for("setup"))
    today = datetime.date.today()
    selected_month = request.args.get("month", default=today.month, type=int)
    selected_year = request.args.get("year", default=today.year, type=int)
    
    available_years = []
    if os.path.exists(PENDAFTARAN_DIR):
        available_years = [int(d) for d in os.listdir(PENDAFTARAN_DIR) if d.isdigit()]
    if not available_years:
        available_years = [selected_year]
    available_years = sorted(available_years)
    
    return render_template("data_penuh.html",
                           selected_month=selected_month,
                           selected_year=selected_year,
                           available_years=available_years,
                           months=MONTH_MAP,
                           klinik_asal=config.get("klinik_asal", "KK BANDAR BARU SELAYANG"),
                           singkatan_klinik=config.get("singkatan_klinik", "KKBBS"))

@app.route("/api/shutdown", methods=["POST"])
def shutdown():
    import signal
    import subprocess
    import platform
    
    def kill_server():
        import time
        time.sleep(1.5)  # Biarkan response sampai ke browser dahulu
        
        # Cuba tutup tetingkap terminal / cmd yang menjalankan proses ini
        try:
            ppid = os.getppid()
            has_console = False
            try:
                # Semak jika python berjalan dalam console aktif
                has_console = sys.stdout and sys.stdout.isatty()
            except Exception:
                has_console = False
                
            if has_console:
                sys_type = platform.system()
                if sys_type == "Darwin":  # macOS
                    try:
                        tty_name = os.ttyname(sys.stdout.fileno())
                        applescript = f'''
                        tell application "Terminal"
                            repeat with w in windows
                                repeat with t in tabs of w
                                    if tty of t is "{tty_name}" then
                                        close w
                                        return
                                    end if
                                end repeat
                            end repeat
                        end tell
                        '''
                        subprocess.Popen(["osascript", "-e", applescript])
                    except Exception:
                        # Fallback jika gagal tty
                        subprocess.Popen(["osascript", "-e", 'tell application "Terminal" to close first window'])
                elif sys_type == "Windows":
                    # Dapatkan nama proses parent
                    try:
                        out = subprocess.check_output(f"tasklist /FI \"PID eq {ppid}\" /NH", shell=True, text=True)
                        if "cmd.exe" in out.lower() or "powershell.exe" in out.lower() or "cmd" in out.lower():
                            subprocess.Popen(f"taskkill /F /PID {ppid}", shell=True)
                    except Exception:
                        pass
        except Exception as e:
            print(f"Ralat semasa menutup terminal: {e}")
            
        # Matikan proses python sendiri
        os.kill(os.getpid(), signal.SIGINT)
        
    threading.Thread(target=kill_server).start()
    return jsonify({"status": "success", "message": "Server sedang ditutup."})


def parse_mykad_py(ic):
    cleaned = "".join(filter(str.isdigit, ic))
    if len(cleaned) != 12:
        return None
        
    dob_part = cleaned[0:6]
    gender_digit = int(cleaned[11])
    
    yy = int(dob_part[0:2])
    mm = int(dob_part[2:4])
    dd = int(dob_part[4:6])
    
    current_year = datetime.date.today().year
    current_year_last_two = current_year % 100
    birth_year = 1900 + yy if yy > current_year_last_two else 2000 + yy
    
    try:
        dob = datetime.date(birth_year, mm, dd)
    except ValueError:
        return None
        
    return {
        "dob": dob,
        "is_male": (gender_digit % 2 != 0)
    }


def calculate_age_from_dob(dob, target_date):
    age = target_date.year - dob.year
    if (target_date.month, target_date.day) < (dob.month, dob.day):
        age -= 1
    return age


@app.route("/patient-list")
def patient_list_page():
    config = load_config()
    if not config.get("is_configured", False):
        return redirect(url_for("setup"))
    return render_template("patient_list.html", klinik_asal=config.get("klinik_asal", "KK RAWANG PERDANA"))


@app.route("/api/patients")
def api_patients():
    start_str = request.args.get("start_date")
    end_str = request.args.get("end_date")
    if not start_str or not end_str:
        return jsonify({"error": "Start date and end date are required"}), 400
        
    try:
        start_date = datetime.datetime.strptime(start_str, "%Y-%m-%d").date()
        end_date = datetime.datetime.strptime(end_str, "%Y-%m-%d").date()
    except ValueError:
        return jsonify({"error": "Invalid date format. Use YYYY-MM-DD"}), 400
        
    delta = end_date - start_date
    dates = [start_date + datetime.timedelta(days=i) for i in range(delta.days + 1)]
    
    file_dates = {}
    for d in dates:
        path = get_excel_path(d)
        if not path or not os.path.exists(path):
            continue
        if path not in file_dates:
            file_dates[path] = []
        file_dates[path].append(d)
        
    patients = []
    
    for path, dates_in_file in file_dates.items():
        try:
            # Load with data_only=True to get evaluated values where possible
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for d in dates_in_file:
                sheetname = str(d.day)
                if sheetname not in wb.sheetnames:
                    continue
                sheet = wb[sheetname]
                
                rows_data = list(sheet.iter_rows(min_row=9, max_row=73, min_col=1, max_col=19, values_only=True))
                for row_idx, row in enumerate(rows_data):
                    xray_no = row[2]
                    if xray_no is not None:
                        id_number = row[3] or ""
                        name = row[4] or ""
                        age = row[5]
                        gender = row[6]
                        
                        # Clean formula string if not evaluated
                        is_mykad = len(str(id_number).replace("-", "")) == 12
                        if str(age).startswith("="):
                            age = None
                        if str(gender).startswith("="):
                            gender = None
                            
                        # Python fallback evaluation for MyKad
                        if is_mykad and (age is None or gender is None):
                            parsed = parse_mykad_py(str(id_number))
                            if parsed:
                                if age is None:
                                    age = calculate_age_from_dob(parsed["dob"], d)
                                if gender is None:
                                    gender = "M" if parsed["is_male"] else "F"
                                    
                        gender_display = "LELAKI"
                        if gender in ["F", "P", "PEREMPUAN", "female", "FEMALE"]:
                            gender_display = "PEREMPUAN"
                        elif gender in ["M", "L", "LELAKI", "male", "MALE"]:
                            gender_display = "LELAKI"
                            
                        patients.append({
                            "tarikh": row[0] or d.strftime("%d-%m-%Y"),
                            "bil_kes": row[1] or (row_idx + 9 - 8),
                            "xray_no": str(xray_no),
                            "id_number": str(id_number),
                            "name": str(name),
                            "age": age,
                            "gender": gender_display,
                            "citizen": row[7] or "",
                            "gov_staff": row[8] or "",
                            "race": row[9] or "",
                            "address": row[10] or "",
                            "exam_type": row[11] or "",
                            "exam_part": row[12] or "",
                            "laterality": row[13] or "",
                            "clinic": row[14] or "",
                            "category": row[15] or "",
                            "lmp": row[16] or "",
                            "cd_filem": row[17] or "",
                            "total_expose": row[18] or 1
                        })
            wb.close()
        except Exception as e:
            print(f"Error reading file {path} for patient list: {e}")
            
    # Sort by X-Ray ascending by default
    patients.sort(key=lambda p: p["xray_no"])
    return jsonify(patients)


def open_browser(port=5005):
    import time
    time.sleep(1.5)
    webbrowser.open(f"http://localhost:{port}")

def show_startup_progress():
    if sys.stdout is None or not hasattr(sys.stdout, 'write'):
        return
    import time
    print("==================================================")
    print("               DAFTAR RADIOLOGI                   ")
    print("               [rdgr.szal.kkbbs]                  ")
    print("==================================================")
    steps = [
        "Checking configuration files...",
        "Validating Excel templates directory...",
        "Initializing web routing system...",
        "Establishing server connections...",
        "Starting local server engine..."
    ]
    for idx, step in enumerate(steps):
        pct = (idx + 1) * 20
        # Draw progress bar
        bar_len = 20
        filled = int(pct * bar_len / 100)
        empty = bar_len - filled
        bar = "#" * filled + "." * empty
        sys.stdout.write(f"\r[{bar}] {pct}% - {step}")
        sys.stdout.flush()
        time.sleep(0.4)
    print("\n--------------------------------------------------")
    print("Daftar Radiologi has started successfully!")
    print("rdgr.szal.kkbbs")
    print("==================================================\n")

if __name__ == "__main__":
    is_frozen = getattr(sys, 'frozen', False)
    if is_frozen or os.environ.get("WERKZEUG_RUN_MAIN") != "true":
        show_startup_progress()
    port = int(os.environ.get("PORT", 5005))
    print(f"Memulakan pelayan di http://localhost:{port}")
    # Buka browser secara automatik apabila server bermula
    if is_frozen or os.environ.get("WERKZEUG_RUN_MAIN") == "true":
        threading.Thread(target=open_browser, args=(port,), daemon=True).start()
    elif not is_frozen:
        # Semasa pembangunan, buka juga browser jika bukan main reloader thread
        threading.Thread(target=open_browser, args=(port,), daemon=True).start()
        
    app.run(host="localhost", port=port, debug=not is_frozen)
