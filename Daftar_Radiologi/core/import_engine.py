import os
import shutil
import datetime
import re
import openpyxl
import csv
from pathlib import Path
from core.config import load_config, PENDAFTARAN_DIR, MONTH_MAP
from core.excel_engine import get_excel_path, add_patient_record
from core.backup_engine import create_zip_backup

TARGET_COLUMNS = [
    {"key": "tarikh", "label": "Tarikh (Date)", "required": True},
    {"key": "no_xray", "label": "No. X-Ray (X-Ray No)", "required": False},
    {"key": "no_ic", "label": "No. KP / Pasport (IC / Passport No)", "required": True},
    {"key": "nama", "label": "Nama Pesakit (Patient Name)", "required": True},
    {"key": "umur", "label": "Umur (Age)", "required": False},
    {"key": "jantina", "label": "Jantina (Gender)", "required": False},
    {"key": "warganegara", "label": "Warganegara (Nationality)", "required": False},
    {"key": "bangsa", "label": "Bangsa (Race)", "required": False},
    {"key": "alamat", "label": "Alamat (Address)", "required": False},
    {"key": "jenis_pemeriksaan", "label": "Jenis Pemeriksaan (Exam Type)", "required": True},
    {"key": "bahagian_pemeriksaan", "label": "Bahagian Pemeriksaan (Exam Part)", "required": True},
    {"key": "laterally", "label": "Laterality (L/R/Bilateral)", "required": False},
    {"key": "klinik_rujukan", "label": "Klinik Rujukan (Referral Clinic)", "required": False},
    {"key": "kategori", "label": "Kategori / Status (Category)", "required": False},
    {"key": "cd_filem", "label": "CD / Filem (Consumables)", "required": False},
    {"key": "juru_xray", "label": "Juru X-Ray (Radiographer)", "required": False}
]

def parse_uploaded_file(file_path):
    """
    Parse an uploaded Excel (.xlsx, .xls) or CSV file.
    Returns dictionary with available sheets, headers, sample rows, and total row count.
    """
    if not os.path.exists(file_path):
        return {"success": False, "message": "Fail tidak wujud."}

    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == ".csv":
        return parse_csv_file(file_path)
    elif ext in [".xlsx", ".xls"]:
        return parse_excel_file(file_path)
    else:
        return {"success": False, "message": "Format fail tidak disokong. Sila guna .xlsx, .xls atau .csv."}

def parse_csv_file(file_path):
    try:
        rows = []
        with open(file_path, "r", encoding="utf-8-sig", errors="ignore") as f:
            reader = csv.reader(f)
            for row in reader:
                if any(str(c).strip() for c in row):
                    rows.append(row)
        
        if not rows:
            return {"success": False, "message": "Fail CSV adalah kosong."}
            
        headers = [str(c).strip() for c in rows[0]]
        sample_rows = [[str(c).strip() for c in r] for r in rows[1:6]]
        
        return {
            "success": True,
            "file_type": "csv",
            "sheets": ["CSV Data"],
            "active_sheet": "CSV Data",
            "headers": headers,
            "sample_rows": sample_rows,
            "total_rows": len(rows) - 1,
            "target_columns": TARGET_COLUMNS
        }
    except Exception as e:
        return {"success": False, "message": f"Ralat membaca CSV: {str(e)}"}

def parse_excel_file(file_path):
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet_names = wb.sheetnames
        
        best_sheet_name = sheet_names[0]
        header_row_idx = 1
        best_headers = []
        sample_rows = []
        max_valid_rows = 0
        
        for sname in sheet_names:
            sheet = wb[sname]
            all_rows = list(sheet.iter_rows(values_only=True))
            
            for idx, r in enumerate(all_rows[:15], start=1):
                if not r:
                    continue
                non_empty = [str(c).strip() for c in r if c is not None and str(c).strip() != ""]
                if len(non_empty) >= 3:
                    candidate_headers = [str(c).strip() if c is not None else f"Column_{i+1}" for i, c in enumerate(r)]
                    data_rows = [
                        [str(c).strip() if c is not None else "" for c in dr]
                        for dr in all_rows[idx:idx+5]
                        if any(c is not None and str(c).strip() != "" for c in dr)
                    ]
                    
                    num_data_rows = len(all_rows) - idx
                    if num_data_rows > max_valid_rows:
                        max_valid_rows = num_data_rows
                        best_sheet_name = sname
                        header_row_idx = idx
                        best_headers = candidate_headers
                        sample_rows = data_rows
                    break
                    
        wb.close()
        
        if not best_headers:
            return {"success": False, "message": "Gagal mengesan header dalam fail Excel."}
            
        return {
            "success": True,
            "file_type": "excel",
            "sheets": sheet_names,
            "active_sheet": best_sheet_name,
            "header_row_idx": header_row_idx,
            "headers": best_headers,
            "sample_rows": sample_rows,
            "total_rows": max_valid_rows,
            "target_columns": TARGET_COLUMNS
        }
    except Exception as e:
        return {"success": False, "message": f"Ralat membaca Excel: {str(e)}"}

def auto_suggest_mapping(headers):
    """
    Cadangkan padanan automatik bagi kolum berasaskan perkataan kata kunci.
    """
    mapping = {}
    
    keywords_map = {
        "tarikh": ["tarikh", "date", "dt"],
        "no_xray": ["xray", "x-ray", "no xray", "no. xray", "film", "no film"],
        "no_ic": ["ic", "kp", "pasport", "passport", "no ic", "nokp", "identity"],
        "nama": ["nama", "name", "pesakit", "patient"],
        "umur": ["umur", "age"],
        "jantina": ["jantina", "gender", "sex"],
        "warganegara": ["warganegara", "nationality", "citizenship"],
        "bangsa": ["bangsa", "race", "ethnicity"],
        "alamat": ["alamat", "address"],
        "jenis_pemeriksaan": ["jenis", "modality", "exam type", "pemeriksaan"],
        "bahagian_pemeriksaan": ["bahagian", "part", "region", "exam part"],
        "laterally": ["laterally", "laterality", "side"],
        "klinik_rujukan": ["klinik", "rujukan", "clinic", "referred"],
        "kategori": ["kategori", "category", "status"],
        "cd_filem": ["cd", "filem", "film", "consumable"],
        "juru_xray": ["juru", "radiographer", "staff"]
    }
    
    for target_key, keywords in keywords_map.items():
        for i, h in enumerate(headers):
            h_clean = h.lower().replace("_", " ").replace(".", "")
            if any(kw in h_clean for kw in keywords):
                mapping[target_key] = i
                break
                
    return mapping

def process_data_migration(file_path, column_mapping, sheet_name=None):
    """
    Melaksanakan migrasi data secara pukal dari fail Excel/CSV ke dalam sistem Buku Daftar.
    Membuat auto-backup sebelum memasukkan data.
    """
    if not os.path.exists(file_path):
        return {"success": False, "message": "Fail data migrasi tidak dijumpai."}

    try:
        backup_res = create_zip_backup()
    except Exception as e_bk:
        print(f"[Import Warning] Auto-backup sebelum import: {e_bk}")

    ext = os.path.splitext(file_path)[1].lower()
    records = []
    
    try:
        if ext == ".csv":
            with open(file_path, "r", encoding="utf-8-sig", errors="ignore") as f:
                reader = list(csv.reader(f))
                if len(reader) > 1:
                    records = reader[1:]
        elif ext in [".xlsx", ".xls"]:
            wb = openpyxl.load_workbook(file_path, data_only=True)
            target_sheet = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active
            all_rows = list(target_sheet.iter_rows(values_only=True))
            
            start_idx = 1
            for idx, r in enumerate(all_rows[:15], start=1):
                if r and len([c for c in r if c is not None and str(c).strip() != ""]) >= 3:
                    start_idx = idx
                    break
                    
            records = all_rows[start_idx:]
            wb.close()
    except Exception as e:
        return {"success": False, "message": f"Gagal membaca fail untuk migrasi: {str(e)}"}

    success_count = 0
    error_count = 0
    error_details = []
    config = load_config()
    default_juru = config.get("default_staff", "")
    
    for row_num, row in enumerate(records, start=1):
        if not row or not any(c is not None and str(c).strip() != "" for c in row):
            continue

        def get_val(key, default=""):
            idx = column_mapping.get(key)
            if idx is not None and idx != "" and int(idx) < len(row):
                val = row[int(idx)]
                return str(val).strip() if val is not None else default
            return default

        tarikh_raw = get_val("tarikh")
        nama = get_val("nama")
        no_ic = get_val("no_ic")
        jenis = get_val("jenis_pemeriksaan")
        bahagian = get_val("bahagian_pemeriksaan")

        if not nama or not no_ic or not jenis or not bahagian:
            error_count += 1
            error_details.append(f"Baris {row_num}: Rekod diabaikan (Nama/IC/Jenis/Bahagian kosong).")
            continue

        formatted_date = datetime.date.today().strftime("%Y-%m-%d")
        if tarikh_raw:
            if isinstance(tarikh_raw, datetime.datetime) or isinstance(tarikh_raw, datetime.date):
                formatted_date = tarikh_raw.strftime("%Y-%m-%d")
            else:
                clean_dt = str(tarikh_raw).split(" ")[0].strip()
                for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%Y/%m/%d"):
                    try:
                        dt_parsed = datetime.datetime.strptime(clean_dt, fmt)
                        formatted_date = dt_parsed.strftime("%Y-%m-%d")
                        break
                    except ValueError:
                        pass

        patient_data = {
            "tarikh": formatted_date,
            "no_xray": get_val("no_xray"),
            "nama": nama.upper(),
            "no_ic": no_ic.upper(),
            "umur": get_val("umur", "-"),
            "jantina": get_val("jantina", "MALE").upper(),
            "warganegara": get_val("warganegara", "MALAYSIA").upper(),
            "bangsa": get_val("bangsa", "MELAYU").upper(),
            "alamat": get_val("alamat", "-").upper(),
            "jenis_pemeriksaan": jenis.upper(),
            "bahagian_pemeriksaan": bahagian.upper(),
            "laterally": get_val("laterally", "-").upper(),
            "klinik_rujukan": get_val("klinik_rujukan", "OPD").upper(),
            "kategori": get_val("kategori", "PESAKIT LUAR").upper(),
            "cd_filem": get_val("cd_filem", "-").upper(),
            "juru_xray": get_val("juru_xray", default_juru).upper()
        }

        try:
            ok, msg, _ = add_patient_record(patient_data)
            if ok:
                success_count += 1
            else:
                error_count += 1
                error_details.append(f"Baris {row_num} ({nama}): {msg}")
        except Exception as e_add:
            error_count += 1
            error_details.append(f"Baris {row_num} ({nama}): {str(e_add)}")

    return {
        "success": True,
        "message": f"Migrasi selesai. {success_count} rekod berjaya dimasukkan.",
        "success_count": success_count,
        "error_count": error_count,
        "error_details": error_details[:20]
    }
