import os
import shutil
import datetime
import gc
import re
import openpyxl
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from pathlib import Path
from core.config import (
    PENDAFTARAN_DIR,
    CONFIG_PATH,
    APP_TEMPLATE_XLSX,
    MONTH_MAP,
    ensure_dirs,
    load_config,
    parse_mykad
)
from core.db_engine import (
    sync_patient_record,
    update_patient_in_db,
    cancel_patient_in_db
)

MONTH_CODES = {
    1: "JAN", 2: "FEB", 3: "MAC", 4: "APR",
    5: "MEI", 6: "JUN", 7: "JUL", 8: "OGOS",
    9: "SEPT", 10: "OKT", 11: "NOV", 12: "DIS"
}

def get_available_years():
    """
    Mengesan senarai tahun yang tersedia daripada fail Excel atau folder pendaftaran.
    """
    today_year = datetime.date.today().year
    years = set([today_year])
    
    if os.path.exists(PENDAFTARAN_DIR):
        for root, dirs, files in os.walk(PENDAFTARAN_DIR):
            for f in files:
                if f.endswith(".xlsx") and not f.startswith("~$"):
                    m = re.search(r'20\d{2}', f)
                    if m:
                        years.add(int(m.group(0)))
            for d in dirs:
                if d.isdigit() and len(d) == 4:
                    years.add(int(d))
                
    return sorted(list(years))

def get_excel_filename_pattern(config, year):
    """
    Menjana nama fail Excel rasmi mengikut konfigurasi fasiliti & modaliti:
    - Mod Hospital (Single Modality): DAFTAR {MODALITY} {SINGKATAN} {YEAR}.xlsx
    - Mod Standard / Multi-Modality: DAFTAR XRAY {SINGKATAN} {YEAR}.xlsx
    """
    singkatan = str(config.get("singkatan_klinik", "")).strip().upper()
    singkatan_part = f" {singkatan}" if singkatan else ""
    facility_type = str(config.get("facility_type", "KK")).strip().upper()
    hospital_scope = str(config.get("hospital_scope", "ALL")).strip().upper()
    single_modality = str(config.get("single_modality", "")).strip().upper()
    
    if facility_type == "HOSPITAL" and hospital_scope == "SINGLE" and single_modality:
        clean_mod = re.sub(r'[^A-Z0-9]', ' ', single_modality).strip()
        clean_mod = re.sub(r'\s+', ' ', clean_mod)
        return f"DAFTAR {clean_mod}{singkatan_part} {year}.xlsx"
    else:
        return f"DAFTAR XRAY{singkatan_part} {year}.xlsx"

def get_excel_path(date_obj):
    """
    Mengambil atau membina jalan fail (.xlsx) pendaftaran setahun.
    Mencari fail sedia ada secara fleksibel atau menyalin dari templat master.
    """
    if isinstance(date_obj, int):
        year = date_obj
    else:
        year = date_obj.year
    
    config = load_config()
    singkatan = str(config.get("singkatan_klinik", "")).strip().upper()
    facility_type = str(config.get("facility_type", "KK")).strip().upper()
    
    os.makedirs(PENDAFTARAN_DIR, exist_ok=True)
    year_dir = os.path.join(PENDAFTARAN_DIR, str(year))
    os.makedirs(year_dir, exist_ok=True)
    
    primary_filename = get_excel_filename_pattern(config, year)
    
    # 1. Senarai calon fail berkeutamaan
    candidates = [
        os.path.join(PENDAFTARAN_DIR, primary_filename),
        os.path.join(year_dir, primary_filename),
        os.path.join(PENDAFTARAN_DIR, f"DAFTAR XRAY {singkatan} {year}.xlsx") if singkatan else None,
        os.path.join(year_dir, f"DAFTAR XRAY {singkatan} {year}.xlsx") if singkatan else None,
        os.path.join(PENDAFTARAN_DIR, f"BUKU DAFTAR XRAY {year}.xlsx"),
        os.path.join(year_dir, f"BUKU DAFTAR XRAY {year}.xlsx"),
        os.path.join(PENDAFTARAN_DIR, f"BUKU DAFTAR XRAY {year} {singkatan}.xlsx") if singkatan else None,
        os.path.join(year_dir, f"BUKU DAFTAR XRAY {year} {singkatan}.xlsx") if singkatan else None
    ]
    candidates = [c for c in candidates if c is not None]
    
    for cand in candidates:
        if os.path.exists(cand):
            return cand

    # 2. Semak sebarang fail pendaftaran sedia ada bagi tahun tersebut
    try:
        for root in [PENDAFTARAN_DIR, year_dir]:
            if os.path.exists(root):
                for f in os.listdir(root):
                    if str(year) in f and ("DAFTAR" in f.upper() or "BUKU" in f.upper() or "PER.SS" in f.upper()) and f.endswith(".xlsx") and not f.startswith("~$"):
                        return os.path.join(root, f)
    except Exception:
        pass

MASTER_TEMPLATE_HEADERS = [
    'TARIKH', 'LMP', 'Bil Kes', 'Nombor X-ray', 'Nombor Kad Pengenalan /Pasport',
    'Nama', 'UMUR', 'JANTINA', 'W/NEGARA', 'KAKITANGAN KERAJAAN', 'BANGSA',
    'ALAMAT', 'JENIS', 'BAHAGIAN', 'LATERALLY', 'KLINIK', 'KATEGORI', 'CD/FILEM',
    'TOTAL EXPOSE', 'TOTAL REJECT', 'FILEM / CD', 'REFERENCE', 'PERSON TAKEN', 'COMMENT', 'DATE'
]

HEADER_COL_WIDTHS = {
    1: 14.0,  # TARIKH
    2: 12.0,  # LMP
    3: 9.0,   # Bil Kes
    4: 15.0,  # Nombor X-ray
    5: 20.0,  # Nombor Kad Pengenalan /Pasport
    6: 38.0,  # Nama
    7: 10.0,  # UMUR
    8: 10.0,  # JANTINA
    9: 12.0,  # W/NEGARA
    10: 16.0, # KAKITANGAN KERAJAAN
    11: 16.0, # BANGSA
    12: 32.0, # ALAMAT
    13: 20.0, # JENIS
    14: 20.0, # BAHAGIAN
    15: 13.0, # LATERALLY
    16: 26.0, # KLINIK
    17: 16.0, # KATEGORI
    18: 15.0, # CD/FILEM
    19: 13.0, # TOTAL EXPOSE
    20: 13.0, # TOTAL REJECT
    21: 15.0, # FILEM / CD
    22: 22.0, # REFERENCE
    23: 22.0, # PERSON TAKEN
    24: 26.0, # COMMENT
    25: 18.0  # DATE
}

def generate_master_template_file(template_path=None, config=None):
    """
    Menjana templat induk Excel (template.xlsx) 12-bulan rasmi KKM secara programmatic:
    - Baris 1 & 2: Background Biru (#2F6EBA) dengan teks Putih
    - Baris 4 (Header): Background Kuning (#FFFF00) dengan teks Hitam Bold
    - 25 Kolum Piawai berdimensi kemas
    """
    if template_path is None:
        template_path = APP_TEMPLATE_XLSX
    if config is None:
        config = load_config()

    os.makedirs(os.path.dirname(template_path), exist_ok=True)
    singkatan = str(config.get("singkatan_klinik", "")).strip().upper()
    facility_name = str(config.get("klinik_asal", "")).strip().upper()
    header_title = f"BUKU DAFTAR RADIOLOGI {singkatan or facility_name or 'KKM'}"

    wb = openpyxl.Workbook()
    # Buang sheet lalai asal
    default_sheet = wb.active

    # Row 1 & 2 Styles (Blue Background #2F6EBA with White text)
    title_fill = PatternFill(start_color='2F6EBA', end_color='2F6EBA', fill_type='solid')
    title_font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
    subtitle_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    title_border = Border(
        left=Side(style='thin', color='204E85'),
        right=Side(style='thin', color='204E85'),
        top=Side(style='thin', color='204E85'),
        bottom=Side(style='thin', color='204E85')
    )
    title_align = Alignment(horizontal='left', vertical='center')

    # Row 4 Styles (Header Yellow Background #FFFF00 with Black bold text)
    header_font = Font(name='Arial', size=9, bold=True, color='000000')
    header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC')
    )
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

    month_names = ["JAN", "FEB", "MAC", "APR", "MEI", "JUN", "JUL", "OGOS", "SEPT", "OKT", "NOV", "DIS"]

    for m_name in month_names:
        sheet = wb.create_sheet(title=m_name)
        sheet.row_dimensions[1].height = 24
        sheet.row_dimensions[2].height = 20
        sheet.row_dimensions[3].height = 6
        sheet.row_dimensions[4].height = 28

        # Baris 1: Background #2F6EBA merentasi Kolum 1 - 25
        for col_idx in range(1, len(MASTER_TEMPLATE_HEADERS) + 1):
            c1 = sheet.cell(row=1, column=col_idx)
            c1.fill = title_fill
            c1.border = title_border
        cell_title = sheet.cell(row=1, column=1)
        cell_title.value = header_title
        cell_title.font = title_font
        cell_title.alignment = title_align
        
        # Baris 2: Background #2F6EBA merentasi Kolum 1 - 25
        for col_idx in range(1, len(MASTER_TEMPLATE_HEADERS) + 1):
            c2 = sheet.cell(row=2, column=col_idx)
            c2.fill = title_fill
            c2.border = title_border
        cell_sub = sheet.cell(row=2, column=1)
        cell_sub.value = "KEMENTERIAN KESIHATAN MALAYSIA — REKOD PENDAFTARAN PESAKIT (PER.SS-RA 101)"
        cell_sub.font = subtitle_font
        cell_sub.alignment = title_align

        # Baris 4: Header 25 Kolum Piawai (#FFFF00)
        for col_idx, h_name in enumerate(MASTER_TEMPLATE_HEADERS, 1):
            cell = sheet.cell(row=4, column=col_idx)
            cell.value = h_name
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = header_align
            
            col_letter = openpyxl.utils.get_column_letter(col_idx)
            w = HEADER_COL_WIDTHS.get(col_idx, 15.0)
            sheet.column_dimensions[col_letter].width = w

    # Padamkan default sheet
    if default_sheet in wb.worksheets:
        wb.remove(default_sheet)

    wb.save(template_path)
    wb.close()
    print(f"[ExcelEngine] Templat Excel induk berjaya dijana di: {template_path}")
    return template_path

def get_excel_path(date_obj):
    """
    Mengambil atau membina jalan fail (.xlsx) pendaftaran setahun.
    Mencari fail sedia ada secara fleksibel atau menyalin dari templat master.
    """
    if isinstance(date_obj, int):
        year = date_obj
    else:
        year = date_obj.year
    
    config = load_config()
    singkatan = str(config.get("singkatan_klinik", "")).strip().upper()
    facility_type = str(config.get("facility_type", "KK")).strip().upper()
    
    os.makedirs(PENDAFTARAN_DIR, exist_ok=True)
    year_dir = os.path.join(PENDAFTARAN_DIR, str(year))
    os.makedirs(year_dir, exist_ok=True)
    
    primary_filename = get_excel_filename_pattern(config, year)
    
    # 1. Senarai calon fail berkeutamaan
    candidates = [
        os.path.join(PENDAFTARAN_DIR, primary_filename),
        os.path.join(year_dir, primary_filename),
        os.path.join(PENDAFTARAN_DIR, f"DAFTAR XRAY {singkatan} {year}.xlsx") if singkatan else None,
        os.path.join(year_dir, f"DAFTAR XRAY {singkatan} {year}.xlsx") if singkatan else None,
        os.path.join(PENDAFTARAN_DIR, f"BUKU DAFTAR XRAY {year}.xlsx"),
        os.path.join(year_dir, f"BUKU DAFTAR XRAY {year}.xlsx"),
        os.path.join(PENDAFTARAN_DIR, f"BUKU DAFTAR XRAY {year} {singkatan}.xlsx") if singkatan else None,
        os.path.join(year_dir, f"BUKU DAFTAR XRAY {year} {singkatan}.xlsx") if singkatan else None
    ]
    candidates = [c for c in candidates if c is not None]
    
    for cand in candidates:
        if os.path.exists(cand):
            return cand

    # 2. Semak sebarang fail pendaftaran sedia ada bagi tahun tersebut
    try:
        for root in [PENDAFTARAN_DIR, year_dir]:
            if os.path.exists(root):
                for f in os.listdir(root):
                    if str(year) in f and ("DAFTAR" in f.upper() or "BUKU" in f.upper() or "PER.SS" in f.upper()) and f.endswith(".xlsx") and not f.startswith("~$"):
                        return os.path.join(root, f)
    except Exception:
        pass

    # 3. Jika fail belum wujud, bina fail baharu daripada templat master
    target_path = candidates[0]
    if not os.path.exists(APP_TEMPLATE_XLSX):
        generate_master_template_file(APP_TEMPLATE_XLSX, config)

    try:
        shutil.copy(APP_TEMPLATE_XLSX, target_path)
        print(f"[ExcelEngine] Fail Pendaftaran {year} dicipta daripada templat: {target_path}")
        sync_single_excel_file(target_path, config, year)
        return target_path
    except Exception as e:
        print(f"[ExcelEngine] Ralat menyalin templat ke {target_path}: {e}")
            
    return target_path

def sync_single_excel_file(filepath, config, year=None):
    """
    Pelarasan tajuk & styling BUKU DAFTAR RADIOLOGI pada fail Excel 12 Bulan.
    - Baris 1 & 2: Background #2F6EBA (Teks Putih)
    - Baris 4: Background #FFFF00 (Teks Hitam Bold)
    """
    if not os.path.exists(filepath):
        return
    try:
        if year is None:
            year = datetime.date.today().year
            
        singkatan = str(config.get("singkatan_klinik", "")).strip().upper()
        facility_name = str(config.get("klinik_asal", "")).strip().upper()
        header_title = f"BUKU DAFTAR RADIOLOGI - {singkatan or facility_name or 'KKM'}"
        
        wb = openpyxl.load_workbook(filepath)
        
        row4_headers = [
            'TARIKH', 'LMP', 'Bil Kes', 'Nombor X-ray', 'Nombor Kad Pengenalan /Pasport',
            'Nama', 'UMUR', 'JANTINA', 'W/NEGARA', 'KAKITANGAN KERAJAAN', 'BANGSA',
            'ALAMAT', 'JENIS', 'BAHAGIAN', 'LATERALLY', 'KLINIK', 'KATEGORI', 'CD/FILEM',
            'TOTAL EXPOSE', 'TOTAL REJECT', 'FILEM / CD', 'REFERENCE', 'PERSON TAKEN', 'COMMENT', 'DATE'
        ]

        title_fill = PatternFill(start_color='2F6EBA', end_color='2F6EBA', fill_type='solid')
        title_font = Font(name='Arial', size=13, bold=True, color='FFFFFF')
        subtitle_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
        title_border = Border(
            left=Side(style='thin', color='204E85'),
            right=Side(style='thin', color='204E85'),
            top=Side(style='thin', color='204E85'),
            bottom=Side(style='thin', color='204E85')
        )
        title_align = Alignment(horizontal='left', vertical='center')

        header_font = Font(name='Arial', size=9, bold=True, color='000000')
        header_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        thin_border = Border(
            left=Side(style='thin', color='CCCCCC'),
            right=Side(style='thin', color='CCCCCC'),
            top=Side(style='thin', color='CCCCCC'),
            bottom=Side(style='thin', color='CCCCCC')
        )
        header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Rename placeholder sheets (e.g. 'JAN {YEAR}' -> 'JAN')
        for i, sname in enumerate(wb.sheetnames):
            month_idx = i + 1
            std_code = MONTH_CODES.get(month_idx, f"M{month_idx}")
            if "{YEAR}" in sname or sname.startswith("SEP") or sname.startswith("JAN"):
                wb[sname].title = std_code
                
        # Update Title row A1:Y1, A2:Y2 & Row 4 Headers
        for sname in wb.sheetnames:
            sheet = wb[sname]
            sheet.row_dimensions[1].height = 24
            sheet.row_dimensions[2].height = 20
            sheet.row_dimensions[4].height = 28

            for col_idx in range(1, len(row4_headers) + 1):
                c1 = sheet.cell(row=1, column=col_idx)
                c1.fill = title_fill
                c1.border = title_border

                c2 = sheet.cell(row=2, column=col_idx)
                c2.fill = title_fill
                c2.border = title_border

            cell_title = sheet.cell(row=1, column=1)
            cell_title.value = header_title
            cell_title.font = title_font
            cell_title.alignment = title_align

            cell_sub = sheet.cell(row=2, column=1)
            cell_sub.value = "KEMENTERIAN KESIHATAN MALAYSIA — REKOD PENDAFTARAN PESAKIT (PER.SS-RA 101)"
            cell_sub.font = subtitle_font
            cell_sub.alignment = title_align

            for col_i, h_val in enumerate(row4_headers, 1):
                c4 = sheet.cell(row=4, column=col_i)
                c4.value = h_val
                c4.font = header_font
                c4.fill = header_fill
                c4.border = thin_border
                c4.alignment = header_align

        wb.save(filepath)
        wb.close()
    except Exception as e:
        print(f"[ExcelEngine] Ralat sync_single_excel_file {filepath}: {e}")

def get_target_sheet(wb, date_obj):
    """
    Mengambil worksheet pendaftaran bulanan mengikut tarikh atau angka bulan.
    """
    if isinstance(date_obj, int):
        month_num = date_obj
        year_num = datetime.date.today().year
    else:
        month_num = date_obj.month
        year_num = date_obj.year
        
    m_code = MONTH_CODES.get(month_num, "JAN")
    alt_code = "SEP" if m_code == "SEPT" else m_code
    folder_name, file_month = MONTH_MAP.get(month_num, ("01_JAN", "1JAN"))
    
    possible_names = [
        m_code,
        alt_code,
        f"{m_code} {year_num}",
        f"{alt_code} {year_num}",
        f"{m_code} {{YEAR}}",
        folder_name,
        file_month,
        str(month_num)
    ]
    
    for name in possible_names:
        if name in wb.sheetnames:
            return wb[name]
            
    for sname in wb.sheetnames:
        if m_code in sname.upper() or alt_code in sname.upper():
            return wb[sname]
            
    if 1 <= month_num <= len(wb.worksheets):
        return wb.worksheets[month_num - 1]

    return wb.worksheets[0]

def get_daily_case_count(date_obj=None):
    """
    Mengira nombor giliran Bil Kes Harian bagi tarikh tertentu.
    Memulangkan (jumlah kes pada tarikh tersebut + 1). Reset ke 1 pada hari baharu.
    """
    if date_obj is None:
        date_obj = datetime.date.today()
    path = get_excel_path(date_obj)
    if not path or not os.path.exists(path):
        return 1
    try:
        formatted_date = date_obj.strftime("%d.%m.%Y")
        formatted_date_slash = date_obj.strftime("%d/%m/%Y")
        formatted_date_dash = date_obj.strftime("%Y-%m-%d")
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = get_target_sheet(wb, date_obj)
        count = 0
        for r in range(5, sheet.max_row + 1):
            val_date = str(sheet.cell(row=r, column=1).value or sheet.cell(row=r, column=2).value or "").strip()
            val_name = str(sheet.cell(row=r, column=6).value or "").strip()
            if val_name and (val_date in (formatted_date, formatted_date_slash, formatted_date_dash)):
                count += 1
        wb.close()
        return count + 1
    except Exception as e:
        print(f"[ExcelEngine V2] Ralat get_daily_case_count: {e}")
        return 1

def get_next_xray_no(date_obj=None):
    """
    Mengesan dan mengembalikan Nombor X-ray seterusnya secara automatik dari helaian bulanan,
    sambil menyemak custom_starting_xray_no jika tetapan offset digunakan.
    """
    if date_obj is None:
        date_obj = datetime.date.today()
        
    config = load_config()
    custom_starting_no = int(config.get("custom_starting_xray_no", 0) or 0)
        
    path = get_excel_path(date_obj)
    if not path or not os.path.exists(path):
        initial_num = max(0, custom_starting_no) + 1
        return f"{initial_num:04d}"
        
    try:
        wb = openpyxl.load_workbook(path, data_only=True)
        sheet = get_target_sheet(wb, date_obj)
        last_num = 0
        
        for row in range(sheet.max_row, 4, -1):
            val = sheet.cell(row=row, column=4).value  # Col D (Nombor X-ray)
            if val is None or str(val).strip() == "":
                val = sheet.cell(row=row, column=3).value
                
            if val is not None:
                val_str = str(val).strip()
                found_nums = [int(n) for n in re.findall(r'\d+', val_str)]
                if found_nums:
                    max_found = max(found_nums)
                    if max_found > last_num:
                        last_num = max_found
                        break

        wb.close()
        final_last = max(last_num, custom_starting_no)
        next_num = final_last + 1
        return f"{next_num:04d}"
    except Exception as e:
        print(f"[ExcelEngine V2] Ralat get_next_xray_no: {e}")
        initial_num = max(0, custom_starting_no) + 1
        return f"{initial_num:04d}"

def add_patient_record(patient_data):
    """
    Menulis rekod pesakit baharu ke fail Excel BUKU DAFTAR XRAY & menyelaraskan ke DB (Dwi-Storan).
    Menyokong multi-examination bagi pesakit yang sama & pengasingan automatik Bilateral.
    Returns (success: bool, message: str, xray_no: str)
    """
    try:
        tarikh_str = patient_data.get("tarikh")
        if tarikh_str:
            try:
                date_obj = datetime.datetime.strptime(tarikh_str, "%Y-%m-%d").date()
            except ValueError:
                date_obj = datetime.date.today()
        else:
            date_obj = datetime.date.today()
            
        formatted_date = date_obj.strftime("%d/%m/%Y")
        
        filepath = get_excel_path(date_obj)
        if not filepath:
            return False, "Gagal membuka fail Excel pendaftaran.", None

        # Dapatkan senarai pemeriksaan (examinations)
        examinations = patient_data.get("examinations")
        if not examinations or not isinstance(examinations, list):
            examinations = [{
                "modality": patient_data.get("modality", "General Radiography"),
                "jenis_pemeriksaan": patient_data.get("jenis_pemeriksaan", "DADA"),
                "bahagian_pemeriksaan": patient_data.get("bahagian_pemeriksaan", "CXR"),
                "lateraliti": patient_data.get("lateraliti") or patient_data.get("laterally", ""),
                "cd_filem": patient_data.get("cd_filem", "CD [1]"),
                "catatan": patient_data.get("catatan", "")
            }]

        base_xray_no = patient_data.get("nombor_xray")
        if not base_xray_no or str(base_xray_no).strip() == "":
            current_num = int(get_next_xray_no(date_obj))
        else:
            found = re.findall(r'\d+', str(base_xray_no))
            if found:
                current_num = int(found[0])
            else:
                current_num = int(get_next_xray_no(date_obj))
            
        config = load_config()
        klinik_asal = config.get("klinik_asal", "")
        
        ic_no = patient_data.get("ic_pasport", "").strip()
        parsed_ic = parse_mykad(ic_no)
        
        umur = patient_data.get("umur")
        jantina = patient_data.get("jantina")
        
        if parsed_ic:
            if not umur or str(umur).strip() == "":
                umur = parsed_ic["age"]
            if not jantina or str(jantina).strip() == "":
                jantina = parsed_ic["gender"]
                
        jantina = "M" if str(jantina).upper() in ["M", "LELAKI", "L"] else "F"
        warganegara = "YA" if str(patient_data.get("warganegara", "YA")).upper() in ["YA", "YES", "Y"] else "TIDAK"
        
        wb = openpyxl.load_workbook(filepath)
        sheet = get_target_sheet(wb, date_obj)
        
        cell_font = Font(name='Arial', size=10)
        thin_side = Side(style='thin', color='CCCCCC')
        cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        align_center = Alignment(horizontal='center', vertical='center')
        align_left = Alignment(horizontal='left', vertical='center')
        left_cols = {6, 12, 13, 14, 16, 22, 23, 24}

        generated_xray_list = []
        synced_exam_records = []
        current_bil_kes = get_daily_case_count(date_obj)

        for exam_idx, exam_item in enumerate(examinations):
            base_bahagian = str(exam_item.get("bahagian_pemeriksaan", "CXR")).strip().upper()
            lat = str(exam_item.get("lateraliti") or exam_item.get("laterally") or "").strip().upper()
            mod = exam_item.get("modality", patient_data.get("modality", "General Radiography"))

            # Pengasingan automatik bagi pemeriksaan Bilateral dalam Excel
            if lat in ["BOTH", "BILATERAL"]:
                num1 = current_num
                num2 = current_num + 1
                current_num += 2

                sub_exams = [
                    {
                        "xray_no": f"{num1:04d}",
                        "lateraliti": "LEFT",
                        "bahagian": f"LEFT {base_bahagian}" if "LEFT" not in base_bahagian and "KIRI" not in base_bahagian else base_bahagian,
                        "bil_kes": current_bil_kes,
                        "modality": mod,
                        "jenis_pemeriksaan": exam_item.get("jenis_pemeriksaan", "DADA"),
                        "cd_filem": exam_item.get("cd_filem", "CD [1]"),
                        "total_expose": exam_item.get("total_expose", patient_data.get("total_expose", 1)),
                        "total_reject": exam_item.get("total_reject", patient_data.get("total_reject", 0)),
                        "catatan": exam_item.get("catatan") or patient_data.get("catatan", "")
                    },
                    {
                        "xray_no": f"{num2:04d}",
                        "lateraliti": "RIGHT",
                        "bahagian": f"RIGHT {base_bahagian}" if "RIGHT" not in base_bahagian and "KANAN" not in base_bahagian else base_bahagian,
                        "bil_kes": current_bil_kes + 1,
                        "modality": mod,
                        "jenis_pemeriksaan": exam_item.get("jenis_pemeriksaan", "DADA"),
                        "cd_filem": exam_item.get("cd_filem", "CD [1]"),
                        "total_expose": exam_item.get("total_expose", patient_data.get("total_expose", 1)),
                        "total_reject": exam_item.get("total_reject", patient_data.get("total_reject", 0)),
                        "catatan": exam_item.get("catatan") or patient_data.get("catatan", "")
                    }
                ]
                current_bil_kes += 2
            else:
                if lat in ["RIGHT", "KANAN"] and "RIGHT" not in base_bahagian and "KANAN" not in base_bahagian:
                    item_bahagian = f"RIGHT {base_bahagian}"
                elif lat in ["LEFT", "KIRI"] and "LEFT" not in base_bahagian and "KIRI" not in base_bahagian:
                    item_bahagian = f"LEFT {base_bahagian}"
                else:
                    item_bahagian = base_bahagian

                sub_exams = [
                    {
                        "xray_no": f"{current_num:04d}",
                        "lateraliti": lat,
                        "bahagian": item_bahagian,
                        "bil_kes": current_bil_kes,
                        "modality": mod,
                        "jenis_pemeriksaan": exam_item.get("jenis_pemeriksaan", "DADA"),
                        "cd_filem": exam_item.get("cd_filem", "CD [1]"),
                        "total_expose": exam_item.get("total_expose", patient_data.get("total_expose", 1)),
                        "total_reject": exam_item.get("total_reject", patient_data.get("total_reject", 0)),
                        "catatan": exam_item.get("catatan") or patient_data.get("catatan", "")
                    }
                ]
                current_num += 1
                current_bil_kes += 1

            for sub in sub_exams:
                generated_xray_list.append(sub["xray_no"])
                synced_exam_records.append(sub)

                # Cari baris kosong pertama
                start_row = 5
                target_row = None
                for r in range(start_row, sheet.max_row + 50):
                    val_name = sheet.cell(row=r, column=6).value  # Col F (Nama)
                    val_ic = sheet.cell(row=r, column=5).value    # Col E (IC)
                    val_xray = sheet.cell(row=r, column=4).value  # Col D (Nombor X-ray)

                    if (val_name is None or str(val_name).strip() == "") and \
                       (val_ic is None or str(val_ic).strip() == "") and \
                       (val_xray is None or str(val_xray).strip() == ""):
                        target_row = r
                        break

                if target_row is None:
                    target_row = sheet.max_row + 1

                exam_catatan = sub.get("catatan") or sub.get("comment") or ""
                bahagian_val = sub.get("bahagian") or sub.get("bahagian_pemeriksaan") or ""
                lateraliti_val = sub.get("lateraliti") or sub.get("laterally") or ""
                filem_cd_val = sub.get("filem_cd") or sub.get("cd_filem") or "CD [1]"

                row_values = [
                    formatted_date,                                              # 1 (A): TARIKH
                    patient_data.get("lmp", ""),                                # 2 (B): LMP
                    str(sub["bil_kes"]),                                         # 3 (C): Bil Kes
                    str(sub["xray_no"]),                                         # 4 (D): Nombor X-ray
                    str(ic_no),                                                  # 5 (E): Nombor Kad Pengenalan /Pasport
                    patient_data.get("nama", "").upper(),                        # 6 (F): Nama
                    int(umur) if str(umur).isdigit() else umur,                  # 7 (G): UMUR
                    jantina,                                                     # 8 (H): JANTINA
                    warganegara,                                                 # 9 (I): W/NEGARA
                    patient_data.get("kakitangan_kerajaan", ""),                # 10 (J): KAKITANGAN KERAJAAN
                    patient_data.get("bangsa", "MELAYU").upper(),                # 11 (K): BANGSA
                    patient_data.get("alamat", "").upper(),                      # 12 (L): ALAMAT
                    sub.get("jenis_pemeriksaan", "CHEST").upper(),               # 13 (M): JENIS
                    str(bahagian_val).upper(),                                   # 14 (N): BAHAGIAN
                    str(lateraliti_val).upper(),                                 # 15 (O): LATERALLY
                    patient_data.get("klinik_rujukan", klinik_asal).upper(),     # 16 (P): KLINIK
                    patient_data.get("kategori", "PESAKIT LUAR").upper(),        # 17 (Q): KATEGORI
                    sub.get("cd_filem", "CD [1]").upper(),                       # 18 (R): CD/FILEM
                    patient_data.get("total_expose", 1),                         # 19 (S): TOTAL EXPOSE
                    patient_data.get("total_reject", 0),                         # 20 (T): TOTAL REJECT
                    str(filem_cd_val).upper(),                                   # 21 (U): FILEM / CD
                    patient_data.get("pegawai_rujukan", "").upper(),              # 22 (V): REFERENCE
                    patient_data.get("operator", "").upper(),                     # 23 (W): PERSON TAKEN
                    str(exam_catatan).upper(),                                   # 24 (X): COMMENT
                    datetime.datetime.now().strftime("%Y-%m-%d %H:%M")           # 25 (Y): DATE
                ]

                row_fill = PatternFill(start_color='F8FAFC', end_color='F8FAFC', fill_type='solid') if (target_row % 2 == 0) else PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')

                for col_idx, val in enumerate(row_values, 1):
                    cell = sheet.cell(row=target_row, column=col_idx)
                    cell.value = val
                    cell.font = cell_font
                    cell.fill = row_fill
                    cell.border = cell_border
                    cell.alignment = align_left if col_idx in left_cols else align_center

        wb.save(filepath)
        wb.close()
        gc.collect()

        # Dwi-Storan: Hantar rekod ke pangkalan data jika diaktifkan (atau giliran offline)
        sync_patient_record(patient_data, synced_exam_records, config)

        # DICOM MWL: Hantar rekod ke antrian DICOM Modality Worklist
        try:
            from core.dicom_engine import add_to_dicom_worklist
            add_to_dicom_worklist(patient_data, synced_exam_records)
        except Exception as e_dicom:
            print(f"[ExcelEngine DICOM Warning] {e_dicom}")

        summary_xray = ", ".join(generated_xray_list)
        return True, f"Pesakit berjaya didaftarkan ({len(generated_xray_list)} pemeriksaan). No. X-ray: {summary_xray}", summary_xray

    except Exception as e:
        print(f"[ExcelEngine V2 ERROR] Gagal mendaftar pesakit: {e}")
        return False, f"Ralat penulisan fail Excel: {str(e)}", None

def edit_patient_record(tarikh_str, xray_no, updated_data):
    """
    Mengemaskini maklumat pesakit sedia ada pada fail Excel dan merefleksikannya ke Pangkalan Data.
    Memelihara susunan baris dan turutan audit.
    """
    try:
        if not tarikh_str:
            date_obj = datetime.date.today()
        else:
            try:
                date_obj = datetime.datetime.strptime(tarikh_str, "%Y-%m-%d").date()
            except ValueError:
                date_obj = datetime.date.today()
                
        filepath = get_excel_path(date_obj)
        if not filepath or not os.path.exists(filepath):
            return False, "Fail Excel pendaftaran tidak dijumpai."
            
        wb = openpyxl.load_workbook(filepath)
        sheet = get_target_sheet(wb, date_obj)
        
        target_row = None
        target_xray_str = str(xray_no).strip().lstrip('0')
        
        for r in range(5, sheet.max_row + 1):
            val_xray = str(sheet.cell(row=r, column=4).value or "").strip()
            clean_xray = val_xray.lstrip('0')
            if clean_xray == target_xray_str or val_xray == str(xray_no).strip():
                target_row = r
                break
                
        if not target_row:
            wb.close()
            return False, f"Rekod dengan No. X-ray {xray_no} tidak dijumpai dalam fail Excel."
            
        # Kemaskini medan-medan pembetulan ralat pendaftaran (25 Kolum Piawai):
        if "tarikh" in updated_data and updated_data["tarikh"]:
            sheet.cell(row=target_row, column=1).value = format_date_ddmmyyyy(updated_data["tarikh"])
        if "lmp" in updated_data:
            sheet.cell(row=target_row, column=2).value = str(updated_data["lmp"]).strip()
        if "ic_pasport" in updated_data:
            sheet.cell(row=target_row, column=5).value = str(updated_data["ic_pasport"]).strip()
        if "nama" in updated_data and updated_data["nama"]:
            sheet.cell(row=target_row, column=6).value = str(updated_data["nama"]).strip().upper()
        if "umur" in updated_data and str(updated_data["umur"]).isdigit():
            sheet.cell(row=target_row, column=7).value = int(updated_data["umur"])
        if "jantina" in updated_data and updated_data["jantina"]:
            sheet.cell(row=target_row, column=8).value = "M" if str(updated_data["jantina"]).upper() in ["M", "LELAKI", "L"] else "F"
        if "warganegara" in updated_data:
            sheet.cell(row=target_row, column=9).value = str(updated_data["warganegara"]).strip().upper()
        if "kakitangan_kerajaan" in updated_data:
            sheet.cell(row=target_row, column=10).value = str(updated_data["kakitangan_kerajaan"]).strip().upper()
        if "bangsa" in updated_data and updated_data["bangsa"]:
            sheet.cell(row=target_row, column=11).value = str(updated_data["bangsa"]).strip().upper()
        if "alamat" in updated_data:
            sheet.cell(row=target_row, column=12).value = str(updated_data["alamat"]).strip().upper()
        if "jenis_pemeriksaan" in updated_data and updated_data["jenis_pemeriksaan"]:
            sheet.cell(row=target_row, column=13).value = str(updated_data["jenis_pemeriksaan"]).strip().upper()
        if "bahagian_pemeriksaan" in updated_data and updated_data["bahagian_pemeriksaan"]:
            sheet.cell(row=target_row, column=14).value = str(updated_data["bahagian_pemeriksaan"]).strip()
        if "lateraliti" in updated_data:
            sheet.cell(row=target_row, column=15).value = str(updated_data["lateraliti"]).strip().upper()
        if "klinik_rujukan" in updated_data and updated_data["klinik_rujukan"]:
            sheet.cell(row=target_row, column=16).value = str(updated_data["klinik_rujukan"]).strip().upper()
        if "kategori" in updated_data:
            sheet.cell(row=target_row, column=17).value = str(updated_data["kategori"]).strip().upper()
        if "cd_filem" in updated_data:
            sheet.cell(row=target_row, column=18).value = str(updated_data["cd_filem"]).strip()
        if "total_expose" in updated_data:
            sheet.cell(row=target_row, column=19).value = updated_data["total_expose"]
        if "total_reject" in updated_data:
            sheet.cell(row=target_row, column=20).value = updated_data["total_reject"]
        if "filem_cd" in updated_data:
            sheet.cell(row=target_row, column=21).value = str(updated_data["filem_cd"]).strip().upper()
        if "pegawai_rujukan" in updated_data:
            sheet.cell(row=target_row, column=22).value = str(updated_data["pegawai_rujukan"]).strip().upper()
        if "operator" in updated_data and updated_data["operator"]:
            sheet.cell(row=target_row, column=23).value = str(updated_data["operator"]).strip().upper()
        if "catatan" in updated_data:
            sheet.cell(row=target_row, column=24).value = str(updated_data["catatan"]).strip().upper()
            
        wb.save(filepath)
        wb.close()
        gc.collect()
        
        # Refleksikan ke Pangkalan Data
        config = load_config()
        update_patient_in_db(xray_no, updated_data, config)

        # DICOM MWL: Kemaskini antrian DICOM Worklist jika berkaitan
        try:
            from core.dicom_engine import update_in_dicom_worklist
            update_in_dicom_worklist(xray_no, updated_data)
        except Exception as e_mwl:
            print(f"[ExcelEngine DICOM Warning] {e_mwl}")
        
        return True, f"Rekod X-Ray {xray_no} berjaya dikemaskini dalam Excel dan Database."
    except Exception as e:
        print(f"[ExcelEngine ERROR] edit_patient_record: {e}")
        return False, f"Ralat semasa mengemaskini rekod: {str(e)}"

def cancel_patient_record(tarikh_str, xray_no, reason, staff_name):
    """
    Pematuhan Kriteria Audit KKM (PER.SS-RA 101):
    Nombor siri X-ray TIDAK BOLEH dipadam sehingga mewujudkan lompang nombor.
    Rekod ditandakan sebagai BATAL pada ruangan Catatan/Comment dan diwarnakan pudar bagi memelihara jejak audit.
    """
    try:
        if not tarikh_str:
            date_obj = datetime.date.today()
        else:
            try:
                date_obj = datetime.datetime.strptime(tarikh_str, "%Y-%m-%d").date()
            except ValueError:
                date_obj = datetime.date.today()
                
        filepath = get_excel_path(date_obj)
        if not filepath or not os.path.exists(filepath):
            return False, "Fail Excel pendaftaran tidak dijumpai."
            
        wb = openpyxl.load_workbook(filepath)
        sheet = get_target_sheet(wb, date_obj)
        
        target_row = None
        target_xray_str = str(xray_no).strip().lstrip('0')
        
        for r in range(5, sheet.max_row + 1):
            val_xray = str(sheet.cell(row=r, column=4).value or "").strip()
            clean_xray = val_xray.lstrip('0')
            if clean_xray == target_xray_str or val_xray == str(xray_no).strip():
                target_row = r
                break
                
        if not target_row:
            wb.close()
            return False, f"Rekod dengan No. X-ray {xray_no} tidak dijumpai."
            
        # 1. Kemaskini ruangan COMMENT (Col 24 / X) dengan status BATAL
        curr_comment = str(sheet.cell(row=target_row, column=24).value or "").strip()
        audit_tag = f"[BATAL: {reason.strip().upper()} OLEH {staff_name.strip().upper()}]"
        new_comment = f"{audit_tag} {curr_comment}".strip()
        sheet.cell(row=target_row, column=24).value = new_comment
        
        # 2. Format visual baris yang dibatalkan
        strike_font = Font(name="Arial", size=9, strike=True, color="6B7280")
        cancel_fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")
        for col_idx in range(1, 26):
            c = sheet.cell(row=target_row, column=col_idx)
            c.font = strike_font
            c.fill = cancel_fill

        wb.save(filepath)
        wb.close()
        gc.collect()
        
        # 3. Refleksikan ke Pangkalan Data
        config = load_config()
        cancel_patient_in_db(xray_no, reason, staff_name, config)

        # DICOM MWL: Buang dari antrian jika kes dibatalkan
        try:
            from core.dicom_engine import remove_from_dicom_worklist
            remove_from_dicom_worklist(xray_no)
        except Exception as e_mwl:
            print(f"[ExcelEngine DICOM Warning] {e_mwl}")
        
        return True, f"Kes X-Ray {xray_no} telah ditandakan BATAL mengikut kriteria audit KKM."
    except Exception as e:
        print(f"[ExcelEngine ERROR] cancel_patient_record: {e}")
        return False, f"Ralat semasa membatalkan rekod: {str(e)}"

def parse_date_val(val):
    if isinstance(val, datetime.date):
        return val
    if isinstance(val, datetime.datetime):
        return val.date()
    if not val:
        return None
    val_str = str(val).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%Y/%m/%d", "%d.%m.%Y"):
        try:
            return datetime.datetime.strptime(val_str, fmt).date()
        except ValueError:
            pass
    return None

def format_date_ddmmyyyy(val):
    """
    Format sebarang nilai tarikh kepada format standard DD/MM/YYYY secara konsisten.
    """
    if not val:
        return ""
    if isinstance(val, (datetime.date, datetime.datetime)):
        d = val.date() if isinstance(val, datetime.datetime) else val
        return d.strftime("%d/%m/%Y")
    val_str = str(val).strip()
    parsed = parse_date_val(val_str)
    if parsed:
        return parsed.strftime("%d/%m/%Y")
    return val_str

def get_patients_list(year=None, month=None, search_query=None, period="month", selected_date_str=None, start_date_str=None, end_date_str=None):
    """
    Membaca senarai pesakit daripada fail Excel sasaran mengikut julat masa (period: 'day', 'week', 'month', 'year', 'custom').
    """
    today = datetime.date.today()
    if year is None:
        year = today.year
    if month is None:
        month = today.month

    target_date = today
    if selected_date_str:
        if "-W" in str(selected_date_str):
            try:
                y_str, w_str = str(selected_date_str).split("-W")
                target_date = datetime.date.fromisocalendar(int(y_str), int(w_str), 1)
            except Exception:
                target_date = today
        else:
            try:
                target_date = datetime.datetime.strptime(selected_date_str, "%Y-%m-%d").date()
            except ValueError:
                target_date = today

    week_start = target_date - datetime.timedelta(days=target_date.weekday())
    week_end = week_start + datetime.timedelta(days=6)

    custom_start_date = None
    custom_end_date = None
    if period == "custom" and start_date_str and end_date_str:
        try:
            custom_start_date = datetime.datetime.strptime(start_date_str, "%Y-%m-%d").date()
            custom_end_date = datetime.datetime.strptime(end_date_str, "%Y-%m-%d").date()
        except ValueError:
            pass

    # Tentukan fail (year, month) yang perlu dibaca
    targets_to_read = []
    if period == "year":
        targets_to_read = [(year, m) for m in range(1, 13)]
    elif period == "week":
        targets_to_read = sorted(list(set([
            (week_start.year, week_start.month),
            (week_end.year, week_end.month)
        ])))
    elif period == "day":
        targets_to_read = [(target_date.year, target_date.month)]
    elif period == "custom" and custom_start_date and custom_end_date:
        cur = datetime.date(custom_start_date.year, custom_start_date.month, 1)
        end_m = datetime.date(custom_end_date.year, custom_end_date.month, 1)
        while cur <= end_m:
            targets_to_read.append((cur.year, cur.month))
            if cur.month == 12:
                cur = datetime.date(cur.year + 1, 1, 1)
            else:
                cur = datetime.date(cur.year, cur.month + 1, 1)
    else:  # month
        targets_to_read = [(year, month)]

    patients = []

    for y_idx, m_idx in targets_to_read:
        try:
            date_obj = datetime.date(y_idx, m_idx, 1)
            filepath = get_excel_path(date_obj)
            if not filepath or not os.path.exists(filepath):
                continue
                
            wb = openpyxl.load_workbook(filepath, data_only=True)
            sheet = get_target_sheet(wb, date_obj)
            
            for r in range(5, sheet.max_row + 1):
                nama = sheet.cell(row=r, column=6).value
                ic = sheet.cell(row=r, column=5).value
                xray_no = sheet.cell(row=r, column=4).value
                
                if not nama and not ic and not xray_no:
                    continue

                # Col 1 (A) is TARIKH in new layout; Col 2 (B) in legacy layout
                cell_date_raw = sheet.cell(row=r, column=1).value
                parsed_d = parse_date_val(cell_date_raw)
                if not parsed_d:
                    parsed_d = parse_date_val(sheet.cell(row=r, column=2).value)
                    if parsed_d:
                        cell_date_raw = sheet.cell(row=r, column=2).value
                        # Legacy fallback: Col 1 is Bil Kes, Col 3 is LMP
                        bil_kes_val = str(sheet.cell(row=r, column=1).value or "")
                        lmp_val = sheet.cell(row=r, column=3).value
                    else:
                        bil_kes_val = str(sheet.cell(row=r, column=3).value or "")
                        lmp_val = sheet.cell(row=r, column=2).value
                else:
                    bil_kes_val = str(sheet.cell(row=r, column=3).value or "")
                    lmp_val = sheet.cell(row=r, column=2).value

                # Penapisan julat masa
                if period == "day":
                    if not parsed_d or parsed_d != target_date:
                        continue
                elif period == "week":
                    if not parsed_d or not (week_start <= parsed_d <= week_end):
                        continue
                elif period == "custom" and custom_start_date and custom_end_date:
                    if not parsed_d or not (custom_start_date <= parsed_d <= custom_end_date):
                        continue
                        
                # Col 24 (X) in new layout; Col 23 (W) in legacy layout
                catatan_val = str(sheet.cell(row=r, column=24).value or sheet.cell(row=r, column=23).value or "")
                is_cancelled = "[BATAL" in catatan_val.upper()
                status = "BATAL" if is_cancelled else "AKTIF"
                
                # Check for 25-col new layout: Col 22=Ref, Col 23=Operator; or legacy Col 21=Ref, Col 22=Operator
                ref_val = str(sheet.cell(row=r, column=22).value or sheet.cell(row=r, column=21).value or "")
                operator_val = str(sheet.cell(row=r, column=23).value or sheet.cell(row=r, column=22).value or "")

                record = {
                    "sheet_name": sheet.title,
                    "row_index": r,
                    "bil_kes": bil_kes_val,
                    "tarikh": format_date_ddmmyyyy(cell_date_raw),
                    "lmp": format_date_ddmmyyyy(lmp_val),
                    "nombor_xray": str(xray_no or ""),
                    "ic_pasport": str(ic or ""),
                    "nama": str(nama or "").upper(),
                    "umur": str(sheet.cell(row=r, column=7).value or ""),
                    "jantina": str(sheet.cell(row=r, column=8).value or ""),
                    "warganegara": str(sheet.cell(row=r, column=9).value or ""),
                    "kakitangan_kerajaan": str(sheet.cell(row=r, column=10).value or ""),
                    "bangsa": str(sheet.cell(row=r, column=11).value or ""),
                    "alamat": str(sheet.cell(row=r, column=12).value or ""),
                    "jenis_pemeriksaan": str(sheet.cell(row=r, column=13).value or ""),
                    "bahagian_pemeriksaan": str(sheet.cell(row=r, column=14).value or ""),
                    "lateraliti": str(sheet.cell(row=r, column=15).value or ""),
                    "laterally": str(sheet.cell(row=r, column=15).value or ""),
                    "klinik_rujukan": str(sheet.cell(row=r, column=16).value or ""),
                    "kategori": str(sheet.cell(row=r, column=17).value or ""),
                    "cd_filem": str(sheet.cell(row=r, column=18).value or ""),
                    "total_expose": sheet.cell(row=r, column=19).value or 1,
                    "total_reject": sheet.cell(row=r, column=20).value or 0,
                    "filem_cd": str(sheet.cell(row=r, column=21).value or ""),
                    "pegawai_rujukan": ref_val,
                    "operator": operator_val,
                    "catatan": catatan_val,
                    "status": status,
                    "is_cancelled": is_cancelled
                }
                
                if search_query:
                    q = str(search_query).lower().strip()
                    combined = f"{record['nama']} {record['ic_pasport']} {record['nombor_xray']} {record['klinik_rujukan']} {record['bahagian_pemeriksaan']}".lower()
                    if q not in combined:
                        continue
                        
                patients.append(record)
                
            wb.close()
        except Exception as e:
            print(f"[ExcelEngine V2 ERROR] Ralat get_patients_list (Bulan {m}): {e}")
            
    return patients

def process_bilateral_record(record):
    """
    Jika rekod mengandungi pendaftaran bilateral, asingkan kepada rekod individu.
    """
    xray_str = str(record.get("nombor_xray", "")).strip()
    lat_str = str(record.get("lateraliti") or record.get("laterally") or "").strip().upper()
    bahagian = str(record.get("bahagian_pemeriksaan", "")).strip()
    
    if lat_str in ["RIGHT", "KANAN"]:
        if "RIGHT" not in bahagian.upper() and "KANAN" not in bahagian.upper():
            record["bahagian_pemeriksaan"] = f"RIGHT {bahagian}".strip()
        record["lateraliti"] = "RIGHT"
        record["laterally"] = "RIGHT"
    elif lat_str in ["LEFT", "KIRI"]:
        if "LEFT" not in bahagian.upper() and "KIRI" not in bahagian.upper():
            record["bahagian_pemeriksaan"] = f"LEFT {bahagian}".strip()
        record["lateraliti"] = "LEFT"
        record["laterally"] = "LEFT"
        
    found_nums = re.findall(r'\d+', xray_str)
    
    if len(found_nums) >= 2:
        rec1 = dict(record)
        rec1["nombor_xray"] = f"{int(found_nums[0]):04d}"
        rec1["lateraliti"] = "LEFT"
        rec1["laterally"] = "LEFT"
        if "LEFT" not in bahagian.upper() and "KIRI" not in bahagian.upper():
            rec1["bahagian_pemeriksaan"] = f"LEFT {bahagian}".strip()
            
        rec2 = dict(record)
        rec2["nombor_xray"] = f"{int(found_nums[1]):04d}"
        rec2["lateraliti"] = "RIGHT"
        rec2["laterally"] = "RIGHT"
        if "RIGHT" not in bahagian.upper() and "KANAN" not in bahagian.upper():
            rec2["bahagian_pemeriksaan"] = f"RIGHT {bahagian}".strip()
            
        return [rec1, rec2]
    
    return [record]

def format_xray_range(xray_list):
    """
    Format senarai nombor X-ray menjadi julat kemas cth: '0001 - 05' atau '0001, 0003'.
    """
    if not xray_list:
        return ""
    nums = []
    for x in xray_list:
        found = re.findall(r'\d+', str(x))
        if found:
            nums.append(int(found[0]))
            
    if not nums:
        return ", ".join(xray_list)
        
    sorted_nums = sorted(list(set(nums)))
    if len(sorted_nums) == 1:
        return f"{sorted_nums[0]:04d}"
        
    is_contiguous = all(sorted_nums[i] == sorted_nums[i-1] + 1 for i in range(1, len(sorted_nums)))
    if is_contiguous:
        start_val = sorted_nums[0]
        end_val = sorted_nums[-1]
        end_str = f"{end_val:02d}" if end_val < 100 else f"{end_val:04d}"
        return f"{start_val:04d} - {end_str}"
    else:
        return ", ".join(f"{n:04d}" for n in sorted_nums)

def get_grouped_patients_list(year=None, month=None, search_query=None, period="month", selected_date_str=None):
    """
    Mengambil dan mengelompokkan rekod pesakit bagi paparan accordion (1 pesakit = 1 baris utama).
    Mengembalikan (grouped_patients, total_cases, total_patients, week_range_text).
    """
    today = datetime.date.today()
    target_date = today
    if selected_date_str:
        if "-W" in str(selected_date_str):
            try:
                y_str, w_str = str(selected_date_str).split("-W")
                target_date = datetime.date.fromisocalendar(int(y_str), int(w_str), 1)
            except Exception:
                target_date = today
        else:
            try:
                target_date = datetime.datetime.strptime(selected_date_str, "%Y-%m-%d").date()
            except ValueError:
                target_date = today

    week_start = target_date - datetime.timedelta(days=target_date.weekday())
    week_end = week_start + datetime.timedelta(days=6)
    week_range_text = f"{week_start.strftime('%d/%m/%Y')} - {week_end.strftime('%d/%m/%Y')}"

    raw_records = get_patients_list(
        year=year,
        month=month,
        search_query=search_query,
        period=period,
        selected_date_str=selected_date_str
    )

    expanded_records = []
    for r in raw_records:
        expanded_records.extend(process_bilateral_record(r))

    groups = {}
    order = []

    for r in expanded_records:
        tarikh = str(r.get("tarikh", "")).strip()
        ic = str(r.get("ic_pasport", "")).strip().upper()
        nama = str(r.get("nama", "")).strip().upper()

        if ic and ic != "-":
            patient_key = f"{tarikh}_{ic}"
        else:
            patient_key = f"{tarikh}_{nama}"

        if patient_key not in groups:
            groups[patient_key] = []
            order.append(patient_key)
        groups[patient_key].append(r)

    grouped_patients = []
    # Mengabaikan rekod yang dibatalkan daripada pengiraan jumlah kes dan pesakit
    active_records = [r for r in expanded_records if not r.get("is_cancelled", False)]
    total_cases = len(active_records)

    active_patient_keys = set()
    for r in active_records:
        t_str = str(r.get("tarikh", "")).strip()
        i_str = str(r.get("ic_pasport", "")).strip().upper()
        n_str = str(r.get("nama", "")).strip().upper()
        pk = f"{t_str}_{i_str}" if (i_str and i_str != "-") else f"{t_str}_{n_str}"
        active_patient_keys.add(pk)

    total_patients = len(active_patient_keys)

    for idx, pkey in enumerate(order, 1):
        items = groups[pkey]
        first = items[0]

        xray_nums = [it.get("nombor_xray", "") for it in items if it.get("nombor_xray")]
        xray_range_display = format_xray_range(xray_nums)

        parts = []
        for it in items:
            p_name = it.get("bahagian_pemeriksaan") or it.get("jenis_pemeriksaan") or ""
            if p_name and p_name not in parts:
                parts.append(p_name)
        bahagian_summary = ", ".join(parts)

        staff_list = []
        for it in items:
            op = it.get("operator", "").strip()
            if op and op not in staff_list:
                staff_list.append(op)
        operator_summary = ", ".join(staff_list)

        catatan_items = [it.get("catatan", "").strip() for it in items if it.get("catatan", "").strip()]
        has_catatan = len(catatan_items) > 0
        has_cancellation = any(it.get("is_cancelled", False) for it in items)

        jantina = first.get("jantina", "").strip()
        bangsa = first.get("bangsa", "").strip()
        jantina_bangsa = f"{jantina} / {bangsa}" if bangsa else jantina

        grouped_patients.append({
            "id": f"patient-{idx}",
            "tarikh": first.get("tarikh", ""),
            "nama": first.get("nama", ""),
            "ic_pasport": first.get("ic_pasport", ""),
            "umur": first.get("umur", ""),
            "jantina": jantina,
            "bangsa": bangsa,
            "jantina_bangsa": jantina_bangsa,
            "klinik_rujukan": first.get("klinik_rujukan", ""),
            "operator": operator_summary,
            "xray_range_display": xray_range_display,
            "bahagian_summary": bahagian_summary,
            "has_catatan": has_catatan,
            "is_cancelled": has_cancellation,
            "status": "BATAL" if has_cancellation else "AKTIF",
            "examinations": items
        })

    return grouped_patients, total_cases, total_patients, week_range_text

def repair_excel_file(year=None):
    """
    Enjin Pemulihan 1-Klik Fail Excel Rosak (1-Click Excel Repair Engine):
    1. Membaca rekod selamat daripada fail Excel sedia ada (rescue mode), DB Sekunder (SQLite/Postgres/MySQL), dan backup zip terkini.
    2. Menyimpan salinan keselamatan fail lama rosak sebagai `.corrupted_{timestamp}.bak`.
    3. Membina semula fail Excel baharu bagi tahun berkenaan berasaskan templat rasmi `template.xlsx` dengan format PER.SS-RA 101 patuh audit.
    4. Memasukkan semula semua rekod pesakit yang dipulihkan secara tersusun dan kemas.
    """
    if not year:
        year = datetime.date.today().year

    config = load_config()
    target_path = get_excel_path(datetime.date(year, 1, 1))

    rescued_records = []
    log_messages = []

    # Step 1: Simpan backup keselamatan fail lama jika wujud
    if os.path.exists(target_path):
        try:
            ts = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
            bak_path = f"{target_path}.corrupted_{ts}.bak"
            shutil.copy2(target_path, bak_path)
            log_messages.append(f"Salinan keselamatan fail lama disimpan: {os.path.basename(bak_path)}")
        except Exception as e:
            log_messages.append(f"Amaran fail backup lama: {e}")

    # Step 2: Cuba rescue daripada fail Excel sedia ada
    try:
        if os.path.exists(target_path):
            existing_recs = get_patients_list(year=year)
            if existing_recs:
                rescued_records.extend(existing_recs)
                log_messages.append(f"Berjaya menyelamatkan {len(existing_recs)} rekod daripada fail Excel sedia ada.")
    except Exception as e:
        log_messages.append(f"Fail Excel sedia ada tidak dapat dibaca sepenuhnya: {e}")

    # Step 3: Rescue daripada DB Sekunder (jika wujud)
    try:
        from core.db_engine import LOCAL_SQLITE_PATH
        if os.path.exists(LOCAL_SQLITE_PATH):
            conn = sqlite3.connect(LOCAL_SQLITE_PATH)
            conn.row_factory = sqlite3.Row
            cursor = conn.cursor()
            cursor.execute("SELECT * FROM patient_examinations WHERE tarikh LIKE ?", (f"{year}-%",))
            db_rows = cursor.fetchall()
            conn.close()
            
            if db_rows:
                existing_keys = {(r.get("tarikh"), str(r.get("nombor_xray")).strip()) for r in rescued_records if r.get("tarikh") and r.get("nombor_xray")}
                added_from_db = 0
                for row in db_rows:
                    dbr = dict(row)
                    key = (dbr.get("tarikh"), str(dbr.get("nombor_xray")).strip())
                    if key not in existing_keys:
                        rescued_records.append(dbr)
                        existing_keys.add(key)
                        added_from_db += 1
                if added_from_db > 0:
                    log_messages.append(f"Berjaya menambah {added_from_db} rekod dipulihkan daripada Pangkalan Data Sekunder.")
    except Exception as e:
        log_messages.append(f"Imbasan DB Sekunder: {e}")

    # Step 4: Re-create fresh Excel file from template.xlsx
    ensure_dirs()
    if not os.path.exists(APP_TEMPLATE_XLSX):
        generate_master_template_file(APP_TEMPLATE_XLSX, config)

    try:
        shutil.copy2(APP_TEMPLATE_XLSX, target_path)
        sync_single_excel_file(target_path, config, year)
        log_messages.append(f"Fail Excel baharu berjaya dibina semula daripada templat rasmi KKM.")
    except Exception as e:
        return False, f"Ralat membina semula templat Excel: {e}", 0

    # Step 5: Tulis semula semua rekod pesakit terselamat
    if rescued_records:
        unique_map = {}
        for r in rescued_records:
            t = r.get("tarikh", "")
            nx = str(r.get("nombor_xray", "")).strip()
            key = f"{t}_{nx}_{r.get('nama','')}"
            if key not in unique_map:
                unique_map[key] = r

        final_records = list(unique_map.values())

        def sort_key(r):
            t = r.get("tarikh", "")
            nx = str(r.get("nombor_xray", ""))
            digits = re.search(r'\d+', nx)
            num = int(digits.group(0)) if digits else 0
            return (t, num)

        final_records.sort(key=sort_key)

        for rec in final_records:
            try:
                add_patient_record(rec)
            except Exception as e:
                print(f"[RepairEngine] Ralat menulis rekod pesakit {rec.get('nama')}: {e}")

        log_messages.append(f"Jumlah {len(final_records)} rekod pesakit telah berjaya dimasukkan semula ke fail Excel baharu.")
        return True, "\n".join(log_messages), len(final_records)

    log_messages.append("Fail Excel dibina semula secara bersih (sedia untuk pendaftaran pesakit baharu).")
    return True, "\n".join(log_messages), 0
